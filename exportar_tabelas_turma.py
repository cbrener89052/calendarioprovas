#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta uma tabela por turma a partir das propostas ja geradas.

Colunas: Disciplina | Professor (sigla - nome) | Dia e tempos | No de tempos

Le  : Horario desenvolvido/Proposta_N_Calendario_Provas_2026_2SEM.xlsx
      siglas/siglas_profs_aux_etc.xlsx  (sigla -> nome do professor)
Grava: Horario desenvolvido/Tabela_Provas_por_Turma_Proposta_N.xlsx
       (uma aba por turma)
"""
import openpyxl, os, re, datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import gerar_calendario as G

BASE = G.BASE
OUT = G.OUT
SEMANA1 = datetime.date(2026, 8, 3)
COLS = {"E": 1, "F": 2, "G": 3, "H": 4, "I": 5}
DIA_NOME = {1: "Segunda", 2: "Terça", 3: "Quarta", 4: "Quinta", 5: "Sexta"}
SIM_COD = re.compile(r"^(AG9|AG10|S\d-\d\d|EX\S+)")


def carregar_siglas():
    """sigla -> 'Nome Sobrenome' da planilha de referencia."""
    wb = openpyxl.load_workbook(
        os.path.join(BASE, "siglas", "siglas_profs_aux_etc.xlsx"), data_only=True)
    ws = wb["Planilha1"]
    d = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sigla, nome, sobrenome = (row + (None, None, None))[:3]
        if not sigla:
            continue
        completo = " ".join(x for x in (nome, sobrenome) if x)
        d[str(sigla).strip()] = completo.strip()
    return d


def nomes_dos_profs(campo, siglas):
    """'BPad/MFo/SMo' ou 'Eth-EFr-Swa' -> 'BPad (Beatriz ...), MFo (Mário ...)'."""
    partes = [p for p in re.split(r"[/\-]", campo) if p.strip()]
    out = []
    for p in partes:
        p = p.strip()
        nome = siglas.get(p)
        if not nome:
            # tenta ignorando maiusculas/minusculas
            for k, v in siglas.items():
                if k.lower() == p.lower():
                    nome = v
                    break
        out.append(f"{p} ({nome})" if nome else p)
    return ", ".join(out)


def data_de(w, d):
    return SEMANA1 + datetime.timedelta(days=7 * (w - 1) + d - 1)


def ler_provas(caminho, turma, siglas):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[turma]
    linhas = []
    for w in range(1, 21):
        r = G.week_row(w)
        for col, d in COLS.items():
            v = ws[f"{col}{r}"].value
            if not v:
                continue
            txt = str(v).strip()
            if ("unterrichtsfrei" in txt or "Fach |" in txt or "Matéria |" in txt
                    or txt.startswith(("2CH", "CC", "Raumwunsch", "U-Stunden"))):
                continue
            partes = [p.strip() for p in txt.split("\n")]
            cabeca = partes[0]
            tempos_txt = partes[-1] if len(partes) > 1 else ""

            if SIM_COD.match(cabeca):
                disc, prof = cabeca, "—"
            elif " - " in cabeca:
                disc, campo = cabeca.split(" - ", 1)
                prof = nomes_dos_profs(campo, siglas)
            else:
                disc, prof = cabeca, "—"

            nums = [int(x) for x in re.findall(r"(\d+)º", tempos_txt)]
            if len(nums) >= 2:
                n_tempos = max(nums) - min(nums) + 1
            elif len(nums) == 1:
                n_tempos = 1
            else:
                n_tempos = ""

            dt = data_de(w, d)
            quando = f"{DIA_NOME[d]}, {dt.strftime('%d/%m/%Y')} — {tempos_txt}"
            linhas.append((dt, disc.strip(), prof, quando, n_tempos))

    linhas.sort()
    # junta os dois dias de um mesmo simulado numa linha so
    juntas, i = [], 0
    while i < len(linhas):
        dt, disc, prof, quando, n = linhas[i]
        if SIM_COD.match(disc) and i + 1 < len(linhas) and linhas[i + 1][1] == disc:
            dt2, _, _, quando2, _ = linhas[i + 1]
            quando = (f"{DIA_NOME[dt.weekday() + 1]}, {dt.strftime('%d/%m')} e "
                      f"{DIA_NOME[dt2.weekday() + 1]}, {dt2.strftime('%d/%m/%Y')}"
                      f" — {quando.split('— ')[-1]}")
            i += 1
        juntas.append((disc, prof, quando, n))
        i += 1
    return juntas


def exportar(proposta, siglas):
    origem = os.path.join(
        OUT, f"Proposta_{proposta}_Calendario_Provas_2026_2SEM.xlsx")
    destino = os.path.join(
        OUT, f"Tabela_Provas_por_Turma_Proposta_{proposta}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cab_fill = PatternFill("solid", fgColor="1F3864")
    cab_font = Font(bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    zebra = PatternFill("solid", fgColor="F2F2F2")

    for turma in G.GRADES:
        ws = wb.create_sheet(turma)
        ws["A1"] = f"Plano de Provas {turma} — 2º semestre 2026 (Proposta {proposta})"
        ws["A1"].font = Font(bold=True, size=13)
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = Alignment(horizontal="center")

        cabecalhos = ["Disciplina", "Professor(es)", "Dia e tempos da prova",
                      "Nº de tempos"]
        for c, t in enumerate(cabecalhos, 1):
            cel = ws.cell(2, c, t)
            cel.fill, cel.font = cab_fill, cab_font
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = borda

        for i, (disc, prof, quando, n) in enumerate(
                ler_provas(origem, turma, siglas), start=3):
            ws.cell(i, 1, disc)
            ws.cell(i, 2, prof)
            ws.cell(i, 3, quando)
            ws.cell(i, 4, n).alignment = Alignment(horizontal="center")
            for c in range(1, 5):
                cel = ws.cell(i, c)
                cel.border = borda
                cel.alignment = Alignment(
                    vertical="center", wrap_text=(c == 2),
                    horizontal="center" if c == 4 else "left")
                if i % 2:
                    cel.fill = zebra

        for col, larg in zip("ABCD", (16, 52, 42, 13)):
            ws.column_dimensions[col].width = larg
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:D{ws.max_row}"

    wb.save(destino)
    return destino


def main():
    siglas = carregar_siglas()
    for p in (1, 2):
        print("Gerado:", exportar(p, siglas))


if __name__ == "__main__":
    main()
