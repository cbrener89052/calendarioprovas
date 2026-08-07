#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera propostas de calendario de provas para as turmas C do 2o semestre 2026.

Le  : Klausurplan_2026_2SEM.xlsx  (aba 10C2 = template com as datas corretas)
Grava: Horario desenvolvido/Proposta_N_...xlsx  (uma aba por turma)
       Horario desenvolvido/Relatorio_trocas_de_tempo.md

As grades das turmas foram extraidas do PDF turmas9a12_2osemestre2026.pdf
por leitura visual e conferidas contra siglas_profs_aux_etc.xlsx.
"""
import openpyxl, shutil, os, random, copy, collections
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# destaque visual para provas que precisaram cruzar o intervalo do recreio
# por falta de qualquer outra alternativa (ver regra na skill)
DESTAQUE_INTERVALO = PatternFill(start_color="FFC000", end_color="FFC000",
                                  fill_type="solid")

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "Klausurplan_2026_2SEM.xlsx")
OUT = os.path.join(BASE, "Horario desenvolvido")

DIAS = {1: "Seg", 2: "Ter", 3: "Qua", 4: "Qui", 5: "Sex"}
COL = {1: "E", 2: "F", 3: "G", 4: "H", 5: "I"}


def week_row(w):
    return 3 + 3 * (w - 1)


# ---------------------------------------------------------------- GRADES
# Notacao compacta: turma -> dia -> {tempo: "disciplina/professor"}
# dia: 1=Seg ... 5=Sex
# TODAS as aulas entram (inclusive as sem prova), porque um tempo sem prova
# ainda pode ser doado para a 2a parte de uma prova.
GRADE_TXT = {
    "9C1": {
        1: {1: "bio/Ale", 2: "geo/Mar", 3: "ing/APa-PaH-Velo", 4: "esp/-", 5: "esp/-",
            6: "GL/Caro-EFr-Eth", 7: "pred/Raf", 9: "apoio/Caro"},
        2: {1: "port/Jana", 2: "DaF/Caro-SGa-EFr", 3: "mat/BrSa", 4: "his/JuLa",
            5: "DaF/Caro-SGa-EFr", 6: "ing/APa-PaH-Velo", 8: "tec/SJu",
            9: "mat/BrSa", 10: "his/JuLa"},
        3: {1: "DaF/Caro-SGa-EFr", 2: "geo/Mar", 3: "bio/Ale", 4: "mat/BrSa",
            5: "mat/BrSa", 6: "GL/Caro-EFr-Eth", 7: "port/Jana", 9: "apoio/FBri"},
        4: {1: "DaF/Caro-SGa-EFr", 2: "DaF/Caro-SGa-EFr", 3: "art/-", 4: "art/-",
            5: "mat/BrSa", 6: "pred/Raf", 7: "fis/VSi", 9: "apoio/SMo"},
        5: {1: "pred/Raf", 2: "ing/APa-PaH-Velo", 3: "ing/APa-PaH-Velo",
            4: "DaF/Caro-SGa-EFr", 5: "GL/Caro-EFr-Eth", 6: "socem/DiS", 7: "qui/Fab"},
    },
    "9C2": {
        1: {1: "pred/Raf", 2: "fis/VSi", 3: "ing/APa-PaH-Velo", 4: "esp/-", 5: "esp/-",
            6: "GL/Caro-EFr-Eth", 7: "his/JuLa", 9: "apoio/Caro"},
        2: {1: "bio/Ale", 2: "DaF/Caro-SGa-EFr", 3: "tec/Tac", 4: "mat/BrSa",
            5: "DaF/Caro-SGa-EFr", 6: "ing/APa-PaH-Velo", 8: "bio/Ale",
            9: "his/JuLa", 10: "mat/BrSa"},
        3: {1: "DaF/Caro-SGa-EFr", 2: "qui/Fab", 3: "mat/BrSa", 4: "port/Jana",
            5: "port/Jana", 6: "GL/Caro-EFr-Eth", 7: "geo/Mar", 9: "apoio/FBri"},
        4: {1: "DaF/Caro-SGa-EFr", 2: "DaF/Caro-SGa-EFr", 3: "art/-", 4: "art/-",
            5: "pred/Raf", 6: "mat/BrSa", 7: "mat/BrSa", 9: "apoio/SMo"},
        5: {1: "socem/DiS", 2: "ing/APa-PaH-Velo", 3: "ing/APa-PaH-Velo",
            4: "DaF/Caro-SGa-EFr", 5: "GL/Caro-EFr-Eth", 6: "geo/Mar", 7: "pred/Raf"},
    },
    "10C1": {
        1: {1: "mat/FBri", 2: "DaF/Eth-EFr-Swa", 3: "pred/BPad", 4: "fis/VSi",
            5: "ing/APa", 6: "tec/SJu", 7: "bio/Lza", 9: "proj/-", 10: "proj/-"},
        2: {1: "DaF/Eth-EFr-Swa", 2: "elet/-", 3: "art/-", 4: "art/-", 5: "port/MFo",
            6: "qui/CAl", 7: "bio/Lza", 9: "gram/SMo", 10: "his/ALu", 11: "mat/BrSa"},
        3: {1: "fil/LAn", 2: "DaF/Eth-EFr-Swa", 3: "GL/EFr-Car-Swa", 4: "ing/APa",
            5: "soc/Kle", 6: "mat/BrSa", 7: "mat/BrSa", 9: "apoio/SGa"},
        4: {1: "fis/VSi", 2: "ing/APa", 3: "ing/APa", 4: "geo/Mlo",
            5: "DaF/Eth-EFr-Swa", 6: "his/ALu", 8: "esp/-", 9: "esp/-",
            10: "pred/BPad", 11: "bio/Lza"},
        5: {1: "GL/EFr-Car-Swa", 2: "geo/Mlo", 3: "DaF/Eth-EFr-Swa", 4: "port/MFo",
            5: "qui/CAl", 6: "mat/FBri", 7: "finan/Mlo"},
    },
    "10C2": {
        1: {1: "port/MFo", 2: "DaF/Eth-EFr-Swa", 3: "mat/FBri", 4: "qui/CAl",
            5: "pred/BPad", 6: "bio/Lza", 7: "fis/VSi", 9: "proj/-", 10: "proj/-"},
        2: {1: "DaF/Eth-EFr-Swa", 2: "elet/-", 3: "art/-", 4: "art/-", 5: "ing/Vir",
            6: "mat/BrSa", 7: "mat/BrSa", 9: "bio/Lza", 10: "bio/Lza", 11: "his/ALu"},
        3: {1: "geo/Mlo", 2: "DaF/Eth-EFr-Swa", 3: "GL/EFr-Car-Swa", 4: "ing/Vir",
            5: "ing/Vir", 6: "mat/FBri", 7: "fil/LAn", 9: "apoio/SGa"},
        4: {1: "geo/Mlo", 2: "port/MFo", 3: "his/ALu", 4: "mat/BrSa",
            5: "DaF/Eth-EFr-Swa", 6: "finan/Mlo", 8: "esp/-", 9: "esp/-",
            10: "fis/VSi", 11: "gram/SMo"},
        5: {1: "GL/EFr-Car-Swa", 2: "ing/Vir", 3: "DaF/Eth-EFr-Swa", 4: "qui/CAl",
            5: "soc/Kle", 6: "tec/Tac", 7: "pred/BPad"},
    },
    "11C1": {
        1: {1: "qui/CAl", 2: "mat/ClaMe", 3: "geo/Mar", 4: "pred/ACo",
            5: "GL/CBu-Swa-SGa", 6: "DaF/CBu-SGa-Swa", 7: "ing/Bea",
            9: "proj/APa", 10: "proj/APa"},
        2: {1: "mat/JJ", 2: "his/Ver", 3: "his/Ver", 4: "DaF/CBu-SGa-Swa", 5: "art/-",
            6: "art/-", 8: "qui/CAl", 9: "bio/Ale", 10: "pred/ACo", 11: "fis/Cadu",
            12: "prve/-", 13: "prve/-"},
        3: {1: "his/Ver", 2: "mat/ClaMe", 3: "apr/-", 4: "GL/CBu-Swa-SGa", 5: "apr/-",
            6: "DaF/CBu-SGa-Swa", 7: "apr/-", 9: "fil/LAn", 10: "bio/Ale",
            11: "mat/ClaMe", 12: "prve/-", 13: "prve/-"},
        4: {1: "fis/Cadu", 2: "fis/Cadu", 3: "mat/JJ", 4: "port/AMu", 5: "ing/Bea",
            6: "ing/Bea", 7: "DaF/CBu-SGa-Swa", 9: "ing/Bea", 10: "esp/-", 11: "esp/-",
            12: "prve/-", 13: "prve/-"},
        5: {1: "apr/-", 2: "soc/Kle", 3: "bio/Ale", 4: "geo/Mar",
            5: "DaF/CBu-SGa-Swa", 6: "gram/Raf", 7: "port/AMu",
            10: "prve/-", 11: "prve/-"},
    },
    "11C2": {
        1: {1: "pred/ACo", 2: "bio/Ale", 3: "mat/ClaMe", 4: "ing/PaH",
            5: "GL/CBu-Swa-SGa", 6: "DaF/CBu-SGa-Swa", 7: "geo/Mar",
            9: "proj/APa", 10: "proj/APa"},
        2: {1: "qui/CAl", 2: "mat/JJ", 3: "qui/CAl", 4: "DaF/CBu-SGa-Swa", 5: "art/-",
            6: "art/-", 8: "fis/Cadu", 9: "pred/ACo", 10: "his/Ver", 11: "his/Ver",
            12: "prve/-", 13: "prve/-"},
        3: {1: "geo/Mar", 2: "his/Ver", 3: "apr/-", 4: "GL/CBu-Swa-SGa", 5: "apr/-",
            6: "DaF/CBu-SGa-Swa", 7: "apr/-", 9: "fis/Cadu", 10: "mat/ClaMe",
            11: "fil/LAn", 12: "prve/-", 13: "prve/-"},
        4: {1: "mat/JJ", 2: "ing/PaH", 3: "bio/Ale", 4: "fis/Cadu", 5: "port/AMu",
            6: "port/AMu", 7: "DaF/CBu-SGa-Swa", 9: "gram/Raf", 10: "esp/-",
            11: "esp/-", 12: "prve/-", 13: "prve/-"},
        5: {1: "apr/-", 2: "bio/Ale", 3: "mat/ClaMe", 4: "soc/Kle",
            5: "DaF/CBu-SGa-Swa", 6: "ing/PaH", 7: "ing/PaH",
            10: "prve/-", 11: "prve/-"},
    },
    "12C1": {
        1: {1: "ing/PaH", 2: "ing/PaH", 3: "bio/Ale", 4: "DaF/CBu-EFr-Eth",
            5: "geo/Mar", 6: "geo/Mar", 7: "qui/CAl", 12: "prve/-", 13: "prve/-"},
        2: {1: "mat/Bre", 2: "bio/Ale", 3: "fis/Cadu", 4: "apr/-", 5: "apr/-",
            6: "GL/CBu-EFr-Eth", 8: "his/Wag", 9: "apr/-", 10: "apr/-",
            11: "port/Deb", 12: "prve/-", 13: "prve/-"},
        3: {1: "bio/Ale", 2: "soc/Kle", 3: "mat/Bre", 4: "fis/Cadu",
            5: "DaF/CBu-EFr-Eth", 6: "qui/Fab", 8: "esp/-", 9: "esp/-",
            10: "fil/LAn", 11: "his/Wag", 12: "prve/-", 13: "prve/-"},
        4: {1: "art/-", 2: "art/-", 3: "pred/AMu", 4: "ing/PaH", 5: "port/Deb",
            6: "mat/JJ", 7: "mat/JJ", 9: "GL/CBu-EFr-Eth", 10: "DaF/CBu-EFr-Eth",
            11: "port/Deb", 12: "prve/-", 13: "prve/-"},
        5: {1: "port/Deb", 2: "geo/Mar", 3: "qui/Fab", 4: "his/Wag", 5: "pred/AMu",
            6: "DaF/CBu-EFr-Eth", 7: "DaF/CBu-EFr-Eth"},
    },
    "12C2": {
        1: {1: "geo/Mar", 2: "qui/CAl", 3: "mat/Bre", 4: "DaF/CBu-EFr-Eth",
            5: "ing/Isb", 6: "ing/Isb", 7: "bio/Ale", 12: "prve/-", 13: "prve/-"},
        2: {1: "fis/Cadu", 2: "port/Deb", 3: "bio/Ale", 4: "apr/-", 5: "apr/-",
            6: "GL/CBu-EFr-Eth", 7: "mat/Bre", 9: "apr/-", 10: "apr/-",
            11: "his/Wag", 12: "prve/-", 13: "prve/-"},
        3: {1: "qui/Fab", 2: "his/Wag", 3: "pred/AMu", 4: "fil/LAn",
            5: "DaF/CBu-EFr-Eth", 6: "soc/Kle", 8: "esp/-", 9: "esp/-",
            10: "port/Deb", 11: "geo/Mar", 12: "prve/-", 13: "prve/-"},
        4: {1: "art/-", 2: "art/-", 3: "fis/Cadu", 4: "mat/JJ", 5: "mat/JJ",
            6: "bio/Ale", 7: "port/Deb", 9: "GL/CBu-EFr-Eth", 10: "DaF/CBu-EFr-Eth",
            11: "ing/Isb", 12: "prve/-", 13: "prve/-"},
        5: {1: "his/Wag", 2: "port/Deb", 3: "geo/Mar", 4: "pred/AMu", 5: "qui/Fab",
            6: "DaF/CBu-EFr-Eth", 7: "DaF/CBu-EFr-Eth"},
    },
}

# disciplinas que NAO geram prova (mas cujo tempo pode ser doado)
SEM_PROVA = {"esp", "art", "tec", "finan", "socem", "apoio", "proj",
             "apr", "elet", "prve"}

GRADES = {
    turma: {(d, t): tuple(v.split("/"))
            for d, tempos in dias.items() for t, v in tempos.items()}
    for turma, dias in GRADE_TXT.items()
}

_GRADES_ANTIGO = {
    "9C1": {
        (1, 1): ("bio", "Ale"), (1, 2): ("geo", "Mar"), (1, 3): ("ing", "APa-PaH-Velo"),
        (1, 6): ("GL", "Caro-EFr-Eth"), (1, 7): ("pred", "Raf"),
        (2, 1): ("port", "Jana"), (2, 2): ("DaF", "Caro-SGa-EFr"), (2, 3): ("mat", "BrSa"),
        (2, 4): ("his", "JuLa"), (2, 5): ("DaF", "Caro-SGa-EFr"), (2, 6): ("ing", "APa-PaH-Velo"),
        (2, 10): ("his", "JuLa"),
        (3, 1): ("DaF", "Caro-SGa-EFr"), (3, 2): ("geo", "Mar"), (3, 3): ("bio", "Ale"),
        (3, 4): ("mat", "BrSa"), (3, 5): ("mat", "BrSa"), (3, 6): ("GL", "Caro-EFr-Eth"),
        (3, 7): ("port", "Jana"),
        (4, 1): ("DaF", "Caro-SGa-EFr"), (4, 2): ("DaF", "Caro-SGa-EFr"), (4, 5): ("mat", "BrSa"),
        (4, 6): ("pred", "Raf"), (4, 7): ("fis", "VSi"),
        (5, 1): ("pred", "Raf"), (5, 2): ("ing", "APa-PaH-Velo"), (5, 4): ("DaF", "Caro-SGa-EFr"),
        (5, 5): ("GL", "Caro-EFr-Eth"), (5, 7): ("qui", "Fab"),
    },
    "9C2": {
        (1, 1): ("pred", "Raf"), (1, 2): ("fis", "VSi"), (1, 3): ("ing", "APa-PaH-Velo"),
        (1, 6): ("GL", "Caro-EFr-Eth"), (1, 7): ("his", "JuLa"),
        (2, 1): ("bio", "Ale"), (2, 2): ("DaF", "Caro-SGa-EFr"), (2, 4): ("mat", "BrSa"),
        (2, 5): ("DaF", "Caro-SGa-EFr"), (2, 6): ("ing", "APa-PaH-Velo"), (2, 8): ("bio", "Ale"),
        (2, 9): ("his", "JuLa"), (2, 10): ("mat", "BrSa"),
        (3, 1): ("DaF", "Caro-SGa-EFr"), (3, 2): ("qui", "Fab"), (3, 3): ("mat", "BrSa"),
        (3, 4): ("port", "Jana"), (3, 6): ("GL", "Caro-EFr-Eth"), (3, 7): ("geo", "Mar"),
        (4, 1): ("DaF", "Caro-SGa-EFr"), (4, 2): ("DaF", "Caro-SGa-EFr"), (4, 5): ("pred", "Raf"),
        (4, 6): ("mat", "BrSa"),
        (5, 2): ("ing", "APa-PaH-Velo"), (5, 4): ("DaF", "Caro-SGa-EFr"),
        (5, 5): ("GL", "Caro-EFr-Eth"), (5, 6): ("geo", "Mar"), (5, 7): ("pred", "Raf"),
    },
    "10C1": {
        (1, 1): ("mat", "FBri"), (1, 2): ("DaF", "Eth-EFr-Swa"), (1, 3): ("pred", "BPad"),
        (1, 4): ("fis", "VSi"), (1, 5): ("ing", "APa"), (1, 7): ("bio", "Lza"),
        (2, 1): ("DaF", "Eth-EFr-Swa"), (2, 5): ("port", "MFo"), (2, 6): ("qui", "CAl"),
        (2, 7): ("bio", "Lza"), (2, 10): ("his", "ALu"), (2, 11): ("mat", "BrSa"),
        (3, 1): ("fil", "LAn"), (3, 2): ("DaF", "Eth-EFr-Swa"), (3, 3): ("GL", "EFr-Car-Swa"),
        (3, 4): ("ing", "APa"), (3, 5): ("soc", "Kle"), (3, 6): ("mat", "BrSa"),
        (3, 7): ("mat", "BrSa"),
        (4, 1): ("fis", "VSi"), (4, 2): ("ing", "APa"), (4, 3): ("ing", "APa"),
        (4, 4): ("geo", "Mlo"), (4, 5): ("DaF", "Eth-EFr-Swa"), (4, 6): ("his", "ALu"),
        (4, 10): ("pred", "BPad"), (4, 11): ("bio", "Lza"),
        (5, 1): ("GL", "EFr-Car-Swa"), (5, 2): ("geo", "Mlo"), (5, 3): ("DaF", "Eth-EFr-Swa"),
        (5, 4): ("port", "MFo"), (5, 5): ("qui", "CAl"), (5, 6): ("mat", "FBri"),
    },
    "10C2": {
        (1, 1): ("port", "MFo"), (1, 2): ("DaF", "Eth-EFr-Swa"), (1, 3): ("mat", "FBri"),
        (1, 4): ("qui", "CAl"), (1, 5): ("pred", "BPad"), (1, 6): ("bio", "Lza"),
        (1, 7): ("fis", "VSi"),
        (2, 1): ("DaF", "Eth-EFr-Swa"), (2, 5): ("ing", "Vir"), (2, 6): ("mat", "BrSa"),
        (2, 7): ("mat", "BrSa"), (2, 9): ("bio", "Lza"), (2, 10): ("bio", "Lza"),
        (2, 11): ("his", "ALu"),
        (3, 1): ("geo", "Mlo"), (3, 2): ("DaF", "Eth-EFr-Swa"), (3, 3): ("GL", "EFr-Car-Swa"),
        (3, 4): ("ing", "Vir"), (3, 5): ("ing", "Vir"), (3, 6): ("mat", "FBri"),
        (3, 7): ("fil", "LAn"),
        (4, 1): ("geo", "Mlo"), (4, 2): ("port", "MFo"), (4, 3): ("his", "ALu"),
        (4, 4): ("mat", "BrSa"), (4, 5): ("DaF", "Eth-EFr-Swa"), (4, 10): ("fis", "VSi"),
        (4, 11): ("port", "SMo"),
        (5, 1): ("GL", "EFr-Car-Swa"), (5, 2): ("ing", "Vir"), (5, 3): ("DaF", "Eth-EFr-Swa"),
        (5, 4): ("qui", "CAl"), (5, 5): ("soc", "Kle"), (5, 7): ("pred", "BPad"),
    },
    "11C1": {
        (1, 1): ("qui", "CAl"), (1, 2): ("mat", "ClaMe"), (1, 3): ("geo", "Mar"),
        (1, 4): ("pred", "ACo"), (1, 5): ("GL", "CBu-Swa-SGa"), (1, 6): ("DaF", "CBu-SGa-Swa"),
        (1, 7): ("ing", "Bea"),
        (2, 1): ("mat", "JJ"), (2, 2): ("his", "Ver"), (2, 3): ("his", "Ver"),
        (2, 4): ("DaF", "CBu-SGa-Swa"), (2, 8): ("qui", "CAl"), (2, 9): ("bio", "Ale"),
        (2, 10): ("pred", "ACo"), (2, 11): ("fis", "Cadu"),
        (3, 1): ("his", "Ver"), (3, 2): ("mat", "ClaMe"), (3, 4): ("GL", "CBu-Swa-SGa"),
        (3, 6): ("DaF", "CBu-SGa-Swa"), (3, 9): ("fil", "LAn"), (3, 10): ("bio", "Ale"),
        (3, 11): ("mat", "ClaMe"),
        (4, 1): ("fis", "Cadu"), (4, 2): ("fis", "Cadu"), (4, 3): ("mat", "JJ"),
        (4, 4): ("port", "AMu"), (4, 5): ("ing", "Bea"), (4, 6): ("ing", "Bea"),
        (4, 7): ("DaF", "CBu-SGa-Swa"), (4, 9): ("ing", "Bea"),
        (5, 2): ("soc", "Kle"), (5, 3): ("bio", "Ale"), (5, 4): ("geo", "Mar"),
        (5, 5): ("DaF", "CBu-SGa-Swa"), (5, 6): ("port", "Raf"), (5, 7): ("port", "AMu"),
    },
    "11C2": {
        (1, 1): ("pred", "ACo"), (1, 2): ("bio", "Ale"), (1, 3): ("mat", "ClaMe"),
        (1, 4): ("ing", "PaH"), (1, 5): ("GL", "CBu-Swa-SGa"), (1, 6): ("DaF", "CBu-SGa-Swa"),
        (1, 7): ("geo", "Mar"),
        (2, 1): ("qui", "CAl"), (2, 2): ("mat", "JJ"), (2, 3): ("qui", "CAl"),
        (2, 4): ("DaF", "CBu-SGa-Swa"), (2, 8): ("fis", "Cadu"), (2, 9): ("pred", "ACo"),
        (2, 10): ("his", "Ver"), (2, 11): ("his", "Ver"),
        (3, 1): ("geo", "Mar"), (3, 2): ("his", "Ver"), (3, 4): ("GL", "CBu-Swa-SGa"),
        (3, 6): ("DaF", "CBu-SGa-Swa"), (3, 9): ("fis", "Cadu"), (3, 10): ("mat", "ClaMe"),
        (3, 11): ("fil", "LAn"),
        (4, 1): ("mat", "JJ"), (4, 2): ("ing", "PaH"), (4, 3): ("bio", "Ale"),
        (4, 4): ("fis", "Cadu"), (4, 5): ("port", "AMu"), (4, 6): ("port", "AMu"),
        (4, 7): ("DaF", "CBu-SGa-Swa"), (4, 9): ("port", "Raf"),
        (5, 2): ("bio", "Ale"), (5, 3): ("mat", "ClaMe"), (5, 4): ("soc", "Kle"),
        (5, 5): ("DaF", "CBu-SGa-Swa"), (5, 6): ("ing", "PaH"), (5, 7): ("ing", "PaH"),
    },
    "12C1": {
        (1, 1): ("ing", "PaH"), (1, 2): ("ing", "PaH"), (1, 3): ("bio", "Ale"),
        (1, 4): ("DaF", "CBu-EFr-Eth"), (1, 5): ("geo", "Mar"), (1, 6): ("geo", "Mar"),
        (1, 7): ("qui", "CAl"),
        (2, 1): ("mat", "Bre"), (2, 2): ("bio", "Ale"), (2, 3): ("fis", "Cadu"),
        (2, 6): ("GL", "CBu-EFr-Eth"), (2, 8): ("his", "Wag"), (2, 11): ("port", "Deb"),
        (3, 1): ("bio", "Ale"), (3, 2): ("soc", "Kle"), (3, 3): ("mat", "Bre"),
        (3, 4): ("fis", "Cadu"), (3, 5): ("DaF", "CBu-EFr-Eth"), (3, 6): ("qui", "Fab"),
        (3, 10): ("fil", "LAn"), (3, 11): ("his", "Wag"),
        (4, 3): ("pred", "AMu"), (4, 4): ("ing", "PaH"), (4, 5): ("port", "Deb"),
        (4, 6): ("mat", "JJ"), (4, 7): ("mat", "JJ"), (4, 9): ("GL", "CBu-EFr-Eth"),
        (4, 11): ("port", "Deb"),
        (5, 1): ("port", "Deb"), (5, 2): ("geo", "Mar"), (5, 3): ("qui", "Fab"),
        (5, 4): ("his", "Wag"), (5, 5): ("pred", "AMu"), (5, 6): ("DaF", "CBu-EFr-Eth"),
        (5, 7): ("DaF", "CBu-EFr-Eth"),
    },
    "12C2": {
        (1, 1): ("geo", "Mar"), (1, 2): ("qui", "CAl"), (1, 3): ("mat", "Bre"),
        (1, 4): ("DaF", "CBu-EFr-Eth"), (1, 5): ("ing", "Isb"), (1, 6): ("ing", "Isb"),
        (1, 7): ("bio", "Ale"),
        (2, 1): ("fis", "Cadu"), (2, 2): ("port", "Deb"), (2, 3): ("bio", "Ale"),
        (2, 6): ("GL", "CBu-EFr-Eth"), (2, 7): ("mat", "Bre"), (2, 11): ("his", "Wag"),
        (3, 1): ("qui", "Fab"), (3, 2): ("his", "Wag"), (3, 3): ("pred", "AMu"),
        (3, 4): ("fil", "LAn"), (3, 5): ("DaF", "CBu-EFr-Eth"), (3, 6): ("soc", "Kle"),
        (3, 10): ("port", "Deb"), (3, 11): ("geo", "Mar"),
        (4, 3): ("fis", "Cadu"), (4, 4): ("mat", "JJ"), (4, 5): ("mat", "JJ"),
        (4, 6): ("bio", "Ale"), (4, 7): ("port", "Deb"), (4, 9): ("GL", "CBu-EFr-Eth"),
        (4, 11): ("ing", "Isb"),
        (5, 1): ("his", "Wag"), (5, 2): ("port", "Deb"), (5, 3): ("geo", "Mar"),
        (5, 4): ("pred", "AMu"), (5, 5): ("qui", "Fab"), (5, 6): ("DaF", "CBu-EFr-Eth"),
        (5, 7): ("DaF", "CBu-EFr-Eth"),
    },
}

# nome legivel da disciplina
NOME = {
    "mat": "Mat", "DaF": "DaF", "GL": "GL", "port": "Port", "plit": "Port",
    "gram": "Gram", "ing": "Ing", "pred": "Redação", "his": "Hist", "geo": "Geo",
    "bio": "Bio", "fis": "Fis", "qui": "Qui", "fil": "Fil", "soc": "Soc",
    "LPLITRED": "LP/LIT/RED",
    "esp": "Ed.Física", "art": "Artes", "tec": "Tec", "finan": "Finanças",
    "socem": "Socioem.", "apoio": "Apoio", "proj": "Projeto", "apr": "Aprof.",
    "elet": "Eletiva", "prve": "Proj.Vestibular",
}

# Turmas 10, 11 e 12: Portugues + Redacao + Gramatica sao UMA prova de
# 3 tempos seguidos, no mesmo dia, com os tres professores.
# Turmas 9: Portugues (2 tempos) e Redacao (2 tempos) sao provas separadas.
COMBINA_PORT = {"port", "pred", "gram"}

# Grupo 1 = Matematica, Alemao (DaF), Portugues e Ingles.
# GL nao entra: e disciplina propria, com prova, mas fora do grupo 1.
GRUPO1 = {"mat", "DaF", "port", "LPLITRED", "ing"}

# Evitar ao maximo provas nos tempos 7 a 11 (o 7o tempo comeca 12h45).
# Nao e proibicao: algumas disciplinas so tem aula nesses horarios.
PRIMEIRO_TEMPO_TARDE = 7         # nao podem coincidir na mesma semana
DOIS_TEMPOS = {"mat", "DaF", "GL", "port", "ing", "pred", "his", "geo", "bio", "fis", "qui"}
UM_TEMPO = {"fil", "soc"}                       # 1 tempo, 1 prova no semestre
NOVE_UM_TEMPO = {"bio", "fis", "qui"}           # nas turmas 9C: 1 tempo, 1 prova no semestre

# ---------------------------------------------------- CALENDARIO / BLOQUEIOS
# semanas: 3..6 = periodo 1 ; 7..16 = periodo 2
P1_SEMANAS = [3, 4, 5, 6]
P2_SEMANAS_10_12 = [7, 8, 9, 10, 12, 13, 14, 15]   # ate 12/11 (quinta da sem 15)
P2_SEMANAS_9_11 = [7, 8, 9, 10, 12, 13, 14, 15, 16]  # ate 21/11 (sem 16, sem sexta 20/11)

BLOQUEIOS = {
    (6, 1),    # 07/09 seg - Independencia
    (14, 1),   # 02/11 seg - Finados
    (16, 5),   # 20/11 sex - Consciencia Negra
    # 12/10 (N. Sra. Aparecida) cai na semana 11, ja vetada por inteiro
}
SEMANA_BLOQUEADA = {11}   # 12/10 a 16/10

LIMITE_DIA = {           # ultima (semana, dia) permitida por grupo de turma
    "10_12": (15, 4),    # 12/11 = quinta da semana 15
    "9_11": (16, 4),     # 20/11 e feriado -> ultimo dia util e quinta 19/11
}

# celulas ja preenchidas no modelo (2CH, CC, unterrichtsfrei...) por turma.
# Preenchido em tempo de execucao por carregar_ocupadas().
OCUPADAS = {}


def carregar_ocupadas():
    """Le do modelo quais (semana, dia) ja tem conteudo, por turma.

    O modelo so traz a aba 10C2 pronta; as demais sao copiadas dela. As
    marcacoes "2CH 10,12" / "CC 9,11" valem so para as series citadas, entao
    sao filtradas por turma.
    """
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["10C2"]
    base = {}
    for w in range(1, 21):
        r = week_row(w)
        for d, c in COL.items():
            v = ws[f"{c}{r}"].value
            if v and str(v).strip():
                base[(w, d)] = str(v).strip()
    for turma in GRADE_TXT:
        serie = "9" if turma.startswith("9") else turma[:2]
        ocup = set()
        for (w, d), txt in base.items():
            t = txt.replace("\n", " ")
            if t.startswith(("AG9", "AG10", "S1-", "S2-", "S3-", "S4-", "EX")):
                continue                      # simulados: tratados a parte
            if t.startswith(("2CH", "CC")):
                series = [s.strip() for s in t.split(None, 1)[-1].split(",")]
                if serie not in series:
                    continue                  # marcacao de outra serie
            if "Fach |" in t or "Matéria |" in t or "Raumwunsch" in t \
               or "U-Stunden" in t or "Solicitação" in t or "Horário" in t:
                continue                      # legenda do modelo
            ocup.add((w, d))
        OCUPADAS[turma] = ocup


SIMULADOS = {
    "9C1": [(9, 5, "AG9", "2º ao 7º tempos")],
    "9C2": [(9, 5, "AG9", "2º ao 7º tempos")],
    "10C1": [(6, 5, "AG10", "2º ao 7º tempos")],
    "10C2": [(6, 5, "AG10", "2º ao 7º tempos")],
    "11C1": [(4, 2, "S3-11", "2º ao 7º tempos"), (4, 3, "S3-11", "2º ao 7º tempos"),
             (13, 1, "S4-11", "2º ao 7º tempos"), (13, 2, "S4-11", "2º ao 7º tempos")],
    "11C2": [(4, 2, "S3-11", "2º ao 7º tempos"), (4, 3, "S3-11", "2º ao 7º tempos"),
             (13, 1, "S4-11", "2º ao 7º tempos"), (13, 2, "S4-11", "2º ao 7º tempos")],
    "12C1": [(7, 3, "S4-12", "2º ao 7º tempos"), (7, 4, "S4-12", "2º ao 7º tempos")],
    "12C2": [(7, 3, "S4-12", "2º ao 7º tempos"), (7, 4, "S4-12", "2º ao 7º tempos")],
}


def grupo_turma(t):
    return "9_11" if t.startswith(("9", "11")) else "10_12"


def semanas_p2(t):
    return P2_SEMANAS_9_11 if grupo_turma(t) == "9_11" else P2_SEMANAS_10_12


def dia_permitido(turma, w, d):
    if w in SEMANA_BLOQUEADA:
        return False
    if (w, d) in BLOQUEIOS:
        return False
    lw, ld = LIMITE_DIA[grupo_turma(turma)]
    if w > lw or (w == lw and d > ld):
        return False
    return True


# ---------------------------------------------------------------- EXAMES
def montar_exames(turma):
    """Retorna lista de exames: (disc, prof, n_tempos, periodo|None)."""
    grade = GRADES[turma]
    discs = {}
    for (d, t), (disc, prof) in grade.items():
        discs.setdefault(disc, set()).add(prof)

    combina = not turma.startswith("9")
    e = []
    if combina:
        profs = sorted({p for d in COMBINA_PORT for p in discs.get(d, ())
                        if p != "-"})
        if profs:
            nome = "/".join(profs)
            e.append(("LPLITRED", nome, 3, 1))
            e.append(("LPLITRED", nome, 3, 2))

    for disc, profs in sorted(discs.items()):
        if disc in SEM_PROVA:
            continue
        if combina and disc in COMBINA_PORT:
            continue                                       # ja entrou combinada
        prof = "/".join(sorted(p for p in profs if p != "-"))
        if disc in UM_TEMPO:
            e.append((disc, prof, 1, None))                # 1 prova no semestre
        elif turma.startswith("9") and disc in NOVE_UM_TEMPO:
            e.append((disc, prof, 1, None))                # 9C: bio/fis/qui 1x, 1 tempo
        elif disc in DOIS_TEMPOS:
            e.append((disc, prof, 2, 1))
            e.append((disc, prof, 2, 2))
    return e


def nao_doadoras(turma):
    """Disciplinas que NAO podem ceder tempo para prova de outra.

    Regra da escola: quem tem um unico tempo de aula na semana (Filosofia,
    Sociologia e afins) nao pode perder esse tempo — seria a aula inteira
    da semana.
    """
    import collections as _c
    cont = _c.Counter(d for (d, _p) in GRADES[turma].values())
    return {d for d, c in cont.items() if c == 1} | UM_TEMPO


# ------------------------------------------------- CARGA HORARIA DO SEMESTRE
# Ultima (semana, dia) com aula normal, por grupo de turma: a data da ultima
# prova de 2a chamada, confirmada pela escola. Depois disso as turmas 10/12
# viajam e as 9/11 encerram o ciclo de avaliacao.
FIM_AULAS = {
    "10_12": (15, 5),   # sexta da semana 15 = 13/11/2026
    "9_11": (17, 5),    # sexta da semana 17 = 27/11/2026
}


def semanas_letivas(turma):
    """Semanas com aula normal: da 1 ate a ultima 2a chamada do grupo,
    fora a semana de conselho de classe (confirmado: sem aula nenhuma)."""
    ultima, _d = FIM_AULAS[grupo_turma(turma)]
    return set(range(1, ultima + 1)) - SEMANA_BLOQUEADA


def aulas_no_semestre(turma, dia):
    """Quantas aulas de um certo dia da semana ocorrem no semestre inteiro
    da turma (um feriado so derruba as aulas do dia da semana dele)."""
    semanas = semanas_letivas(turma)
    faltas = sum(1 for (sw, sd) in BLOQUEIOS if sd == dia and sw in semanas)
    return len(semanas) - faltas


def programadas_no_semestre(turma):
    """{(disc, prof): nº de aulas programadas no semestre} da turma."""
    import collections as _c
    cont = _c.Counter()
    for (d, _t), (disc, prof) in GRADES[turma].items():
        cont[(disc, prof)] += aulas_no_semestre(turma, d)
    return cont


def aulas_semanais(turma):
    """{(disc, prof): nº de aulas por semana} da turma."""
    import collections as _c
    return _c.Counter(GRADES[turma].values())


# ----------------------------------------------------- LIMITES DE CESSAO
# Parametros da Proposta 3 (pedido da coordenacao). Ver a secao "Limites de
# cessao de aula" na skill.
META_CESSAO_POUCAS_AULAS = 2     # disciplinas de 2 ou 3 aulas semanais
# Excecao confirmada pela escola: estas podem ceder 3 no semestre.
META_CESSAO_EXCECAO = {"his": 3, "geo": 3, "GL": 3}
LIMITE_PCT_CESSAO = 0.10         # alvo de cessao
# Teto duro: "nenhuma disciplina deve ceder mais que 11%". A folga entre o
# alvo e o teto e o que permite as 3 cessoes das disciplinas de 2 aulas
# semanais acima (3 de 28 aulas = 10,7%).
TETO_PCT_CESSAO = 0.11


def meta_cessao(disc, n_sem):
    """Teto de cessoes pela regra 1, ja com as excecoes por disciplina."""
    if n_sem > 3:
        return None                  # sem teto proprio: vale so o percentual
    return META_CESSAO_EXCECAO.get(disc, META_CESSAO_POUCAS_AULAS)


def limite_cessao(turma, disc, prof, folga=0):
    """Maximo de aulas que (disc, prof) pode ceder no semestre naquela turma.

    Combina as regras 1, 2 e 5 da Proposta 3, sempre pela mais restritiva:
      - 1 aula por semana: nao cede nada (regra 2, ja valia antes);
      - 2 ou 3 aulas por semana: no maximo 2 cessoes -- 3 para Historia,
        Geografia e GL (regra 1 + excecao da escola);
      - todas: no maximo 11% das aulas programadas no semestre (regra 5).
    'folga' afrouxa o limite em N unidades — usado so quando a busca
    estrita nao fecha, para achar a melhor solucao possivel.
    """
    n_sem = aulas_semanais(turma).get((disc, prof), 0)
    if n_sem <= 1 or disc in UM_TEMPO:
        return 0
    prog = programadas_no_semestre(turma).get((disc, prof), 0)
    lim = int(prog * TETO_PCT_CESSAO)
    meta = meta_cessao(disc, n_sem)
    if meta is not None:
        lim = min(lim, meta)
    return lim + folga


def _tarde(t, n):
    """True se a prova encostar nos tempos 7 a 11 (a partir das 12h45)."""
    return (t + n - 1) >= PRIMEIRO_TEMPO_TARDE


INTERVALOS = (3, 5)   # recreio depois do 3o tempo e depois do 5o tempo


def cruza_intervalo(t, n):
    """True se o bloco [t, t+n-1] atravessar o recreio (depois do 3o ou do
    5o tempo). Restricao de maior prioridade: so aceita como ultimo
    recurso (ver 'Regras de distribuicao das provas' na skill)."""
    fim = t + n - 1
    return any(t <= b and fim >= b + 1 for b in INTERVALOS)


def slots_da_disciplina(turma, disc, n_tempos):
    """(dia, tempo_inicial, doador) onde cabe a prova.

    Para provas de 2 tempos o par pode ser (t, t+1) ou (t-1, t): o tempo
    emprestado tanto pode vir depois quanto antes do tempo da disciplina.
    """
    grade = GRADES[turma]
    out = []

    veto = nao_doadoras(turma)

    if disc == "LPLITRED":
        # trio de tempos consecutivos; quanto mais tempos ja forem de
        # Portugues/Redacao/Gramatica, melhor (menos tempo emprestado)
        cand = []
        for (d, t) in sorted(grade):
            trio = [grade.get((d, t + k)) for k in range(3)]
            if any(x is None for x in trio):
                continue
            proprios = sum(1 for x in trio if x[0] in COMBINA_PORT)
            if proprios == 0:
                continue
            doadores = [x for x in trio if x[0] not in COMBINA_PORT]
            if any(x[0] in veto for x in doadores):
                continue                      # disciplina de 1 tempo nao doa
            cand.append((cruza_intervalo(t, 3), _tarde(t, 3), -proprios, d, t,
                         tuple(doadores)))
        cand.sort()
        return [(d, t, doad or None) for (_, _, _, d, t, doad) in cand]

    for (d, t), (dd, _) in sorted(grade.items()):
        if dd != disc:
            continue
        if n_tempos == 1:
            out.append((d, t, None))
            continue
        prox = grade.get((d, t + 1))
        if prox is not None and (prox[0] == disc or prox[0] not in veto):
            out.append((d, t, None if prox[0] == disc else prox))
        ant = grade.get((d, t - 1))
        if ant is not None and ant[0] != disc and ant[0] not in veto:
            out.append((d, t - 1, ant))
    # remove duplicatas e poe os horarios sem cruzar intervalo e da manha
    # na frente
    vistos, unicos = set(), []
    for s in sorted(out, key=lambda s: (cruza_intervalo(s[1], n_tempos),
                                        _tarde(s[1], n_tempos), s[0], s[1])):
        k = (s[0], s[1])
        if k not in vistos:
            vistos.add(k)
            unicos.append(s)
    return unicos


MAX_NOS = 60000

# Orcamento de nos quando os limites de cessao da Proposta 3 estao ativos.
# Cada no fica bem mais caro (checagem de teto/contato/vespera de prova):
# medido, ~600 nos/s contra dezenas de milhares sem os limites. Como uma
# solucao viavel e achada em menos de 3 mil nos, um teto alto so faz os
# degraus inviaveis da escada de relaxamento custarem minutos cada um.
MAX_NOS_CESSAO = 5000


def escada(n_exames, cessoes):
    """Degraus (max_intervalo, max_tarde, max_g1) a tentar, em ordem.

    Sem limites de cessao: escada exaustiva, do mais estrito ao mais
    folgado — e o que as Propostas 1 e 2 sempre usaram.

    Com os limites de cessao ligados, cada no da busca fica muito mais
    caro e a escada exaustiva (~1300 degraus) nao termina em tempo util.
    Como a ordenacao das opcoes dentro da busca ja prefere, nessa ordem,
    nao cruzar o intervalo e nao usar os tempos 7-11, uma escada curta
    chega praticamente na mesma qualidade em poucos segundos.
    """
    if cessoes is None:
        return [(mi, mt, mg)
                for mi in range(0, n_exames + 1)
                for mt in range(0, n_exames + 1)
                for mg in (1, 2, 3)]
    n = n_exames + 1
    # o degrau com max_g1=2 e obrigatorio: sem ele a busca pula direto para
    # 3 disciplinas do grupo 1 na mesma semana, o que a regra proibe
    return [(0, 0, 1), (0, n, 1), (0, n, 2), (n, n, 2), (n, n, 3)]

# Provas com dia/tempo forcados por exigencia externa (turma, disc, periodo)
# -> (semana, dia). Ex.: 1a prova de Ingles da 10C2 tem que cair em
# 01/09/2026 (terca, semana 5) -- pedido da coordenacao.
FORCAR_DATA = {
    ("10C2", "ing", 1): (5, 2),
}


def par_g1_permitido(atuais, disc):
    """Excecao confirmada pela escola: duas provas do grupo 1 na mesma
    semana so sao aceitaveis quando uma delas for Ingles."""
    g1 = [a for a in atuais if eh_grupo1(a)] + [disc]
    return any(x == "ing" or x == "__forcado_ing__" for x in g1)


def eh_grupo1(nome):
    """True para disciplina do grupo 1 ou para o marcador de reserva de uma
    prova do grupo 1 com data forcada (ver FORCAR_DATA)."""
    if nome in GRUPO1:
        return True
    if nome.startswith("__forcado_") and nome.endswith("__"):
        return nome[len("__forcado_"):-2] in GRUPO1
    return False


def doador_discs(doador):
    """Nomes de disciplina dos tempos doados (0, 1 ou varios)."""
    if not doador:
        return set()
    if isinstance(doador[0], tuple):
        return {dd for (dd, _dp) in doador}
    return {doador[0]}


def familia_de(disc):
    """Disciplinas da grade que compoem a prova (LP/LIT/RED = as tres)."""
    return set(COMBINA_PORT) if disc == "LPLITRED" else {disc}


class Cessoes:
    """Estado das cessoes de tempo de uma turma durante o backtracking.

    Implementa as regras 1 a 5 da Proposta 3. Todo metodo de checagem e
    puro (nao altera estado); 'aplicar'/'desfazer' andam em par com o
    empilha/desempilha da busca.
    """

    # Cache por turma de quantas aulas cada (disc, prof) tem em cada semana
    # (descontando feriados). E o denominador da regra 3 e nao muda durante
    # a busca — recalcular no laco interno inviabilizava o tempo de solucao.
    _CACHE_AULAS = {}

    def __init__(self, turma, folga=0, inicial=None, regra3=True, regra4=True):
        self.turma = turma
        self.folga = folga
        self.regra3 = regra3
        self.regra4 = regra4
        self.slots = collections.defaultdict(set)   # (disc,prof) -> {(w,d,t)}
        self.exames = collections.defaultdict(set)  # disc da grade -> {semana}
        self.ultima_semana = FIM_AULAS[grupo_turma(turma)][0]
        # (disc,prof) -> {semana: nº de aulas dessa semana}
        if turma not in Cessoes._CACHE_AULAS:
            porsem = collections.defaultdict(lambda: collections.Counter())
            for (d, t), chave in GRADES[turma].items():
                for w in range(1, self.ultima_semana + 1):
                    if w in SEMANA_BLOQUEADA or (w, d) in BLOQUEIOS:
                        continue
                    porsem[chave][w] += 1
            Cessoes._CACHE_AULAS[turma] = {k: dict(v) for k, v in porsem.items()}
        self._aulas_sem = Cessoes._CACHE_AULAS[turma]
        # (disc,prof) -> {semana: nº de aulas ja cedidas nela}
        self.cedidas_sem = collections.defaultdict(lambda: collections.Counter())
        # disc da grade -> {semanas em que cedeu} (indice para a regra 4)
        self.semanas_cedidas = collections.defaultdict(collections.Counter)
        # Semanas em que cada disciplina NAO pode ceder porque ja tem data
        # de prova exigida pela coordenacao (FORCAR_DATA): pela regra 4, uma
        # cessao na semana da prova ou na anterior tornaria essa data
        # impossivel -- e a data e inegociavel.
        self.proibidas = collections.defaultdict(set)
        for (turma2, disc2, _per2), (wf, _df) in FORCAR_DATA.items():
            if turma2 == turma:
                self.proibidas[disc2] |= {wf - 1, wf}
        for (chave, slots) in (inicial or {}).items():
            for (w, d, t) in slots:
                self.slots[chave].add((w, d, t))
                self.cedidas_sem[chave][w] += 1
                self.semanas_cedidas[chave[0]][w] += 1

    # ---------------------------------------------------------- consultas
    def _sem_contato(self, chave, w, extra_w=None):
        """A disciplina fica sem NENHUMA aula com a turma na semana w?

        extra_w: semana de uma cessao candidata ainda nao aplicada."""
        if w < 1 or w > self.ultima_semana:
            return False                     # fora do semestre letivo
        if w in SEMANA_BLOQUEADA:
            return True                      # conselho de classe: sem aula
        total = self._aulas_sem.get(chave, {}).get(w, 0)
        if total == 0:
            return True
        cedidas = self.cedidas_sem[chave][w] + (1 if extra_w == w else 0)
        return cedidas >= total

    def _pares_sem_contato(self, chave, extra_w=None):
        """Semanas w em que w e w+1 ficam ambas sem contato com a turma.

        Varre o semestre inteiro — usar so fora do laco de busca (ex.: no
        verificador). Dentro da busca, use _viola_regra3, que e O(1)."""
        out = set()
        anterior = self._sem_contato(chave, 1, extra_w)
        for w in range(1, self.ultima_semana + 1):
            atual = self._sem_contato(chave, w + 1, extra_w)
            if anterior and atual:
                out.add(w)
            anterior = atual
        return out

    def _viola_regra3(self, chave, w, quantas):
        """Ceder 'quantas' aulas de chave na semana w deixaria a disciplina
        sem contato com a turma em duas semanas seguidas?

        So a semana w muda de status, entao basta olhar w-1, w e w+1. Conta
        apenas o que a cessao CAUSA: se w ja estava sem contato por outro
        motivo (feriado no unico dia da disciplina), nao foi a cessao."""
        if self._sem_contato(chave, w):
            return False                     # ja era assim antes da cessao
        self.cedidas_sem[chave][w] += quantas
        try:
            if not self._sem_contato(chave, w):
                return False                 # ainda sobra aula na semana
        finally:
            self.cedidas_sem[chave][w] -= quantas
        return self._sem_contato(chave, w - 1) or self._sem_contato(chave, w + 1)

    def pode_ceder_bloco(self, w, d, t, n, disc):
        """Regras 1 a 5 para o bloco inteiro de uma prova candidata.

        Avalia o bloco de uma vez porque um mesmo doador pode ceder mais de
        um tempo na mesma prova — checar tempo a tempo subestimaria o
        consumo do teto.
        """
        if not self.pode_alocar_exame(disc, w):
            return False
        fam = familia_de(disc)
        pedidos = collections.defaultdict(list)
        for k in range(n):
            dono = GRADES[self.turma].get((d, t + k))
            if not dono or dono[0] in fam:
                continue
            pedidos[dono].append((w, d, t + k))
        for (chave, novos) in pedidos.items():
            disc_d, prof_d = chave
            teto = limite_cessao(self.turma, disc_d, prof_d, self.folga)
            if len(self.slots[chave]) + len(novos) > teto:
                return False
            if self.regra4:
                for we in self.exames.get(disc_d, ()):
                    if we in (w, w + 1):
                        return False
            # as datas exigidas pela coordenacao sao protegidas mesmo com a
            # regra 4 relaxada: sao inegociaveis
            if w in self.proibidas.get(disc_d, ()):
                return False
            # regra 3, com todos os tempos do bloco contados juntos (podem
            # ser 2 ou 3 aulas do mesmo doador no mesmo dia)
            if self.regra3 and self._viola_regra3(chave, w, len(novos)):
                return False
        return True

    def pode_alocar_exame(self, disc, w):
        """Regra 4, pelo lado da prova: nenhuma das disciplinas que compoem
        esta prova pode ter cedido aula na semana da prova ou na anterior
        (a aula anterior a prova e justamente a de revisao)."""
        if not self.regra4:
            return True
        for cod in familia_de(disc):
            semanas = self.semanas_cedidas.get(cod)
            if semanas and (w in semanas or (w - 1) in semanas):
                return False
        return True

    # ------------------------------------------------------------ mutacao
    def aplicar(self, w, d, t, n, disc, doador):
        """Registra a prova e as cessoes que ela provoca. Devolve o que for
        preciso para desfazer."""
        registradas = []
        for k in range(n):
            dono = GRADES[self.turma].get((d, t + k))
            if not dono or dono[0] in familia_de(disc):
                continue
            slot = (w, d, t + k)
            if slot in self.slots[dono]:
                continue
            self.slots[dono].add(slot)
            self.cedidas_sem[dono][w] += 1
            self.semanas_cedidas[dono[0]][w] += 1
            registradas.append((dono, slot))
        # a semana da prova entra sob CADA disciplina que a compoe: numa
        # prova combinada (LP/LIT/RED) quem cede tempo e Português,
        # Redação ou Gramática, e e por esse nome que a regra 4 consulta
        novos_exames = [c for c in familia_de(disc) if w not in self.exames[c]]
        for c in novos_exames:
            self.exames[c].add(w)
        return (registradas, novos_exames, w)

    def desfazer(self, marca):
        registradas, novos_exames, w = marca
        for (dono, slot) in registradas:
            self.slots[dono].discard(slot)
            self.cedidas_sem[dono][slot[0]] -= 1
            # zerar tem que APAGAR a chave: a regra 4 testa 'semana in
            # semanas_cedidas', e um contador zerado continuaria presente
            contagem = self.semanas_cedidas[dono[0]]
            contagem[slot[0]] -= 1
            if contagem[slot[0]] <= 0:
                del contagem[slot[0]]
        for c in novos_exames:
            self.exames[c].discard(w)

    def resumo(self):
        """{(disc, prof): nº de aulas cedidas} — para o relatorio."""
        return {k: len(v) for k, v in self.slots.items() if v}

    # ---------------------------------------------------------- snapshot
    # Uma tentativa de busca que estoura MAX_NOS aborta por excecao, sem
    # desempilhar — por isso cada tentativa roda sobre um clone e so o
    # clone vencedor e adotado.
    def clone(self):
        novo = Cessoes(self.turma, self.folga, regra3=self.regra3,
                       regra4=self.regra4)
        novo.slots = collections.defaultdict(set,
                                             {k: set(v) for k, v in self.slots.items()})
        novo.exames = collections.defaultdict(set,
                                              {k: set(v) for k, v in self.exames.items()})
        novo.cedidas_sem = collections.defaultdict(
            lambda: collections.Counter(),
            {k: collections.Counter(v) for k, v in self.cedidas_sem.items()})
        novo.semanas_cedidas = collections.defaultdict(
            collections.Counter,
            {k: collections.Counter(v) for k, v in self.semanas_cedidas.items()})
        return novo

    def adotar(self, outro):
        self.slots = outro.slots
        self.exames = outro.exames
        self.cedidas_sem = outro.cedidas_sem
        self.semanas_cedidas = outro.semanas_cedidas


def _tentar(turma, seed, max_g1, max_tarde, max_intervalo,
            preocupado_dia=None, pre_por_semana=None, excluir=None,
            cessoes=None):
    """Backtracking sob tres folgas controladas, da de maior para a de
    menor prioridade (quem chama itera max_intervalo por fora de
    max_tarde, que fica por fora de max_g1):

    max_intervalo : quantas provas podem cruzar o intervalo do recreio
                    (3o/4o ou 5o/6o tempos) -- maior prioridade de todas
    max_g1        : quantas disciplinas do grupo 1 podem cair na mesma semana
    max_tarde     : quantas provas da turma podem usar os tempos 7 a 11
    preocupado_dia, pre_por_semana : (w,d) e disciplinas ja comprometidas
        por provas conjuntas com a turma irma (ver resolver_par) -- contam
        para os limites desta turma mas nao sao realocadas aqui.
    excluir : nomes de disciplina ja resolvidos por resolver_par (nao
        entram nesta busca, so contam via preocupado_dia/pre_por_semana).
    cessoes : objeto Cessoes com os limites de cessao de aula da Proposta 3.
        None (Propostas 1 e 2) = sem essas restricoes.
    """
    rnd = random.Random(seed)
    exames = [e for e in montar_exames(turma) if e[0] not in (excluir or ())]
    rnd.shuffle(exames)
    exames.sort(key=lambda e: len(slots_da_disciplina(turma, e[0], e[2])))

    ocupado_dia = set(preocupado_dia or ())        # (w, d)
    por_semana = {w: list(v) for w, v in (pre_por_semana or {}).items()}
    for (w, d, cod, _lab) in SIMULADOS[turma]:
        ocupado_dia.add((w, d))
        atuais = por_semana.setdefault(w, [])
        if cod not in atuais:
            atuais.append(cod)
    # dias reservados so para segurar o lugar de uma prova com data forcada
    # (ver resolver_par) -- aqui e onde essa prova de fato vai ser alocada,
    # entao o dia precisa ficar livre para ela.
    for (turma2, _disc2, _per2), (w, d) in FORCAR_DATA.items():
        if turma2 == turma:
            ocupado_dia.discard((w, d))

    resultado = []
    nos = [0]
    tarde = [0]
    intervalo_cnt = [0]

    def cabe(w, d, disc, marcador=None):
        if not dia_permitido(turma, w, d):
            return False
        if (w, d) in OCUPADAS.get(turma, ()):
            return False
        if (w, d) in ocupado_dia:
            return False
        atuais = por_semana.get(w, [])
        # o marcador de reserva (ver FORCAR_DATA) representa esta mesma
        # prova -- nao conta como uma avaliacao extra na semana
        n_atuais = len(atuais) - (1 if marcador and marcador in atuais else 0)
        if n_atuais >= 3:
            return False
        if disc in atuais:
            return False
        if disc in GRUPO1:
            if sum(1 for a in atuais if eh_grupo1(a)) >= max_g1:
                return False
            if sum(1 for a in atuais if eh_grupo1(a)) >= 1 and \
               not par_g1_permitido(atuais, disc):
                return False
        return True

    SLOTS = {}
    for (disc, prof, n, per) in exames:
        SLOTS[(disc, n)] = slots_da_disciplina(turma, disc, n)

    def semanas_do(per):
        if per == 1:
            return P1_SEMANAS
        if per == 2:
            return semanas_p2(turma)
        # exame unico no semestre: prefere o periodo 2, que tem mais folga
        return list(semanas_p2(turma)) + list(P1_SEMANAS)

    def opcoes(ex):
        disc, prof, n, per = ex
        forc = FORCAR_DATA.get((turma, disc, per))
        marcador = f"__forcado_{disc}__" if forc else None
        out = []
        semanas = [forc[0]] if forc else semanas_do(per)
        for w in semanas:
            for (d, t, doador) in SLOTS[(disc, n)]:
                if forc and d != forc[1]:
                    continue
                if cruza_intervalo(t, n) and intervalo_cnt[0] >= max_intervalo:
                    continue
                if _tarde(t, n) and tarde[0] >= max_tarde:
                    continue
                if cessoes is not None and \
                   not cessoes.pode_ceder_bloco(w, d, t, n, disc):
                    continue
                if cabe(w, d, disc, marcador):
                    out.append((w, d, t, doador))
        return out

    def donor_hist():
        hist = {}
        for (_w, _d, _t, _n, dc, _p, doad) in resultado:
            hist.setdefault(dc, set()).update(doador_discs(doad))
        return hist

    def bt(restantes):
        nos[0] += 1
        if nos[0] > (MAX_NOS_CESSAO if cessoes is not None else MAX_NOS):
            raise TimeoutError
        if not restantes:
            return True
        # MRV: resolve primeiro o exame com menos opcoes
        melhor, melhor_ops = None, None
        for ex in restantes:
            ops = opcoes(ex)
            if not ops:
                return False
            if melhor_ops is None or len(ops) < len(melhor_ops):
                melhor, melhor_ops = ex, ops
                if len(ops) == 1:
                    break
        disc, prof, n, per = melhor
        rnd.shuffle(melhor_ops)
        # Ordem de preferencia (da maior para a menor prioridade):
        # 1) nunca cruzar o intervalo do recreio
        # 2) evitar os tempos 7-11
        # 3) espalhar (semanas menos cheias)
        # 4) alternar de qual lado vem o tempo emprestado, quando a mesma
        #    disciplina ja tomou tempo do mesmo doador numa prova anterior
        #    do semestre (desempate, nunca piora os criterios acima)
        # 5) slot melhor ranqueado em SLOTS (LP/LIT/RED = menos tempo
        #    emprestado)
        hist = donor_hist().get(disc, set())
        ordem = {(d, t): i for i, (d, t, _) in enumerate(SLOTS[(disc, n)])}
        melhor_ops.sort(key=lambda o: (
            cruza_intervalo(o[2], n),
            _tarde(o[2], n),
            len(por_semana.get(o[0], [])),
            1 if doador_discs(o[3]) & hist else 0,
            ordem.get((o[1], o[2]), 99),
        ))
        resto = [e for e in restantes if e is not melhor]
        for (w, d, t, doador) in melhor_ops:
            cruza = cruza_intervalo(t, n)
            if cruza and intervalo_cnt[0] >= max_intervalo:
                continue
            eh_tarde = _tarde(t, n)
            if eh_tarde and tarde[0] >= max_tarde:
                continue
            # revalida: o estado de cessoes mudou desde que opcoes() rodou
            if cessoes is not None and \
               not cessoes.pode_ceder_bloco(w, d, t, n, disc):
                continue
            if cruza:
                intervalo_cnt[0] += 1
            if eh_tarde:
                tarde[0] += 1
            ocupado_dia.add((w, d))
            marcador = f"__forcado_{disc}__" if FORCAR_DATA.get((turma, disc, per)) else None
            semana_atual = por_semana.setdefault(w, [])
            tinha_marcador = marcador is not None and marcador in semana_atual
            if tinha_marcador:
                semana_atual.remove(marcador)
            semana_atual.append(disc)
            resultado.append((w, d, t, n, disc, prof, doador))
            marca = cessoes.aplicar(w, d, t, n, disc, doador) \
                if cessoes is not None else None
            if bt(resto):
                return True
            if marca is not None:
                cessoes.desfazer(marca)
            resultado.pop()
            por_semana[w].remove(disc)
            if tinha_marcador:
                por_semana[w].append(marcador)
            ocupado_dia.discard((w, d))
            if eh_tarde:
                tarde[0] -= 1
            if cruza:
                intervalo_cnt[0] -= 1
        return False

    try:
        return bt(list(exames)), resultado
    except TimeoutError:
        return False, []


def resolver(turma, seed, preocupado_dia=None, pre_por_semana=None, excluir=None,
             cessoes=None):
    """Procura a melhor solucao, afrouxando as folgas na ordem de
    prioridade: primeiro tenta nunca cruzar o intervalo do recreio; so
    depois de esgotar isso passa a tentar reduzir provas nos tempos 7-11;
    e so por ultimo aceita sobreposicao do grupo 1."""
    n_exames = len([e for e in montar_exames(turma) if e[0] not in (excluir or ())])
    for (max_intervalo, max_tarde, max_g1) in escada(n_exames, cessoes):
        tentativa = cessoes.clone() if cessoes is not None else None
        ok, res = _tentar(turma, seed + 17 * max_tarde + 131 * max_intervalo,
                          max_g1, max_tarde, max_intervalo,
                          preocupado_dia, pre_por_semana, excluir,
                          tentativa)
        if ok:
            if cessoes is not None:
                cessoes.adotar(tentativa)
            usadas = sum(1 for (_w, _d, t, n, *_r) in res if _tarde(t, n))
            cruzadas = sum(1 for (_w, _d, t, n, *_r) in res
                           if cruza_intervalo(t, n))
            return ok, res, max_g1, usadas, cruzadas
    return False, [], 3, 0, 0


# ------------------------------------------------- PROVAS ENTRE TURMAS IRMAS
# Quando a mesma pessoa leciona uma disciplina nas duas turmas de uma serie
# (ex.: 10C1/10C2), a prova tem que ser aplicada simultaneamente -- o
# professor nao pode estar em dois lugares ao mesmo tempo. Isso vale mesmo
# quando a disciplina ja e "grupo paralelo" que abrange as duas turmas
# (Alemao/DaF, GL, Ingles em algumas series): mesmo la, a SEMANA da prova
# ainda e uma escolha unica que tem que ser igual nas duas abas. Ver a
# regra "Disciplina com professor comum entre turmas irmas" na skill.
PARES_IRMAS = [("9C1", "9C2"), ("10C1", "10C2"), ("11C1", "11C2"), ("12C1", "12C2")]


def _discs_de(disc):
    return COMBINA_PORT if disc == "LPLITRED" else {disc}


def profs_do_disc(turma, disc):
    alvo = _discs_de(disc)
    return {p for (dd, p) in GRADES[turma].values() if dd in alvo}


def posicoes_do_disc(turma, disc):
    alvo = _discs_de(disc)
    return {(d, t) for (d, t), (dd, _p) in GRADES[turma].items() if dd in alvo}


def classificar_par(a, b):
    """disc -> ('combinada'|'coordenar', profs) para as disciplinas com
    professor comum entre as turmas irmas a/b. 'combinada': mesmo
    dia/tempo hoje na grade (ja e grupo paralelo entre as duas turmas,
    so falta escolher a semana). 'coordenar': mesmo professor, tempos
    diferentes em cada turma (precisa achar um dia/tempo em comum)."""
    exa = {d: p for (d, p, _n, _per) in montar_exames(a)}
    exb = {d: p for (d, p, _n, _per) in montar_exames(b)}
    out = {}
    for disc in sorted(set(exa) & set(exb)):
        if exa[disc] != exb[disc]:
            continue                       # professores diferentes -> independente
        pa, pb = posicoes_do_disc(a, disc), posicoes_do_disc(b, disc)
        tipo = "combinada" if pa == pb else "coordenar"
        out[disc] = (tipo, exa[disc])
    return out


def bloco(turma, e_propria, d, t, n, veto):
    """Tempos [t, t+n-1] do dia d na turma: lista de n itens, cada um None
    (tempo proprio da disciplina) ou (disc, prof) de quem doa aquele
    tempo. None (o retorno inteiro) se algum tempo nao existir na grade
    (dia sem aula ou almoco) ou for de disciplina que nao pode doar."""
    out = []
    for k in range(n):
        cell = GRADES[turma].get((d, t + k))
        if cell is None:
            return None
        dd, dp = cell
        if e_propria(dd):
            out.append(None)
        else:
            if dd in veto:
                return None
            out.append((dd, dp))
    return out


def doador_de_bloco(bl):
    doados = [x for x in bl if x is not None]
    if not doados:
        return None
    if len(doados) == 1:
        return doados[0]
    return tuple(doados)


def _tentar_par(a, b, seed, max_g1, max_tarde, max_intervalo, cessoes=None):
    """Como _tentar(), mas decide de uma vez as provas de professor comum
    entre as turmas irmas a/b, no mesmo dia e tempo nas duas.

    cessoes : {turma: Cessoes} com os limites da Proposta 3, ou None."""
    comuns = classificar_par(a, b)
    exames = []
    for (d, p, n, per) in montar_exames(a):
        if d in comuns:
            exames.append((d, p, n, per))
    if not exames:
        return True, [], [], set(), {}, set(), {}

    # Disciplinas que NAO sao resolvidas aqui (professores diferentes nas
    # duas turmas -- na pratica o Ingles das turmas 10/11/12): a prova
    # delas e alocada depois, por turma. Se estas provas coordenadas
    # tomarem os tempos dessas disciplinas, a regra 4 ("nao ceder na
    # semana da propria prova nem na anterior") pode deixa-las sem
    # nenhuma semana livre. Evitar toma-las como doadoras.
    pendentes = {d for (d, _p, _n, _per) in montar_exames(a) if d not in comuns}
    for turma_x in (a, b):
        pendentes |= {d for (d, _p, _n, _per) in montar_exames(turma_x)
                      if d not in comuns}

    def custo_pendente(bl):
        return sum(1 for x in bl if x is not None and x[0] in pendentes)

    rnd = random.Random(seed)
    rnd.shuffle(exames)

    veto_a, veto_b = nao_doadoras(a), nao_doadoras(b)

    ocupado_a, ocupado_b = set(), set()
    por_semana_a, por_semana_b = {}, {}
    for (turma, ocup, psem) in ((a, ocupado_a, por_semana_a), (b, ocupado_b, por_semana_b)):
        for (w, d, cod, _lab) in SIMULADOS[turma]:
            ocup.add((w, d))
            atuais = psem.setdefault(w, [])
            if cod not in atuais:
                atuais.append(cod)
        # reserva os dias de provas com data forcada (ex.: Ingles 10C2 em
        # 01/09) para que as provas conjuntas desta secao nao os ocupem.
        # Usa um marcador (nao o nome real da disciplina) para nao colidir
        # com a checagem "disciplina ja nessa semana" quando a prova real
        # for alocada depois, no resolver() da propria turma.
        for (turma2, disc2, _per2), (w, d) in FORCAR_DATA.items():
            if turma2 != turma:
                continue
            ocup.add((w, d))
            marcador = f"__forcado_{disc2}__"
            atuais = psem.setdefault(w, [])
            if marcador not in atuais:
                atuais.append(marcador)

    resultado_a, resultado_b = [], []
    nos = [0]
    tarde = [0]
    intervalo_cnt = [0]

    def cabe(turma, ocup, psem, w, d, disc):
        if not dia_permitido(turma, w, d):
            return False
        if (w, d) in OCUPADAS.get(turma, ()):
            return False
        if (w, d) in ocup:
            return False
        atuais = psem.get(w, [])
        if len(atuais) >= 3:
            return False
        if disc in atuais:
            return False
        if disc in GRUPO1:
            if sum(1 for x in atuais if eh_grupo1(x)) >= max_g1:
                return False
            if sum(1 for x in atuais if eh_grupo1(x)) >= 1 and \
               not par_g1_permitido(atuais, disc):
                return False
        return True

    def semanas_do(per):
        if per == 1:
            return P1_SEMANAS
        if per == 2:
            return semanas_p2(a)           # a e b sao do mesmo grupo de turma
        return list(semanas_p2(a)) + list(P1_SEMANAS)

    # Blocos (dia, tempo) estruturalmente validos de cada prova, calculados
    # UMA vez: nao dependem da semana nem do estado da busca. Sem isso, cada
    # no reenumerava 5 dias x 11 tempos para cada prova restante, o que
    # inviabilizava o tempo de solucao com os limites de cessao ligados.
    BLOCOS = {}
    for (_disc, _prof, _n, _per) in exames:
        if (_disc, _n) in BLOCOS:
            continue
        e_propria = (lambda dd: dd in COMBINA_PORT) if _disc == "LPLITRED" \
            else (lambda dd, _x=_disc: dd == _x)
        cand = []
        for d in range(1, 6):
            for t in range(1, 12):
                bl_a = bloco(a, e_propria, d, t, _n, veto_a)
                if bl_a is None:
                    continue
                bl_b = bloco(b, e_propria, d, t, _n, veto_b)
                if bl_b is None:
                    continue
                if all(x is not None for x in bl_a) and \
                   all(x is not None for x in bl_b):
                    continue        # nenhuma das duas turmas tem tempo proprio ali
                custo = sum(x is not None for x in bl_a) + \
                    sum(x is not None for x in bl_b)
                cand.append((d, t, bl_a, bl_b, custo))
        BLOCOS[(_disc, _n)] = cand

    def opcoes(ex):
        disc, prof, n, per = ex
        out = []
        for w in semanas_do(per):
            for (d, t, bl_a, bl_b, custo) in BLOCOS[(disc, n)]:
                if cruza_intervalo(t, n) and intervalo_cnt[0] >= max_intervalo:
                    continue
                if _tarde(t, n) and tarde[0] >= max_tarde:
                    continue
                if not cabe(a, ocupado_a, por_semana_a, w, d, disc):
                    continue
                if not cabe(b, ocupado_b, por_semana_b, w, d, disc):
                    continue
                if cessoes is not None and not all(
                        cessoes[tu].pode_ceder_bloco(w, d, t, n, disc)
                        for tu in (a, b)):
                    continue
                out.append((w, d, t, bl_a, bl_b, custo))
        return out

    def bt(restantes):
        nos[0] += 1
        if nos[0] > (MAX_NOS_CESSAO if cessoes is not None else MAX_NOS):
            raise TimeoutError
        if not restantes:
            return True
        # MRV: resolve primeiro a prova com menos opcoes. Com os limites de
        # cessao ligados, avaliar TODAS as provas restantes a cada no e o
        # grosso do custo, entao para na primeira suficientemente restrita.
        corte = 1 if cessoes is None else 3
        melhor, melhor_ops = None, None
        for ex in restantes:
            ops = opcoes(ex)
            if not ops:
                return False
            if melhor_ops is None or len(ops) < len(melhor_ops):
                melhor, melhor_ops = ex, ops
                if len(ops) <= corte:
                    break
        disc, prof, n, per = melhor
        rnd.shuffle(melhor_ops)
        melhor_ops.sort(key=lambda o: (
            cruza_intervalo(o[2], n),
            _tarde(o[2], n),
            # nao consumir tempo de quem ainda precisa alocar a propria
            # prova depois (ver 'pendentes'): so desempate, nunca piora as
            # regras acima
            custo_pendente(o[3]) + custo_pendente(o[4]),
            len(por_semana_a.get(o[0], [])) + len(por_semana_b.get(o[0], [])),
            o[5],
        ))
        resto = [e for e in restantes if e is not melhor]
        for (w, d, t, bl_a, bl_b, _custo) in melhor_ops:
            # revalida: o estado de cessoes mudou desde que opcoes() rodou
            if cessoes is not None and not all(
                    cessoes[tu].pode_ceder_bloco(w, d, t, n, disc)
                    for tu in (a, b)):
                continue
            cruza = cruza_intervalo(t, n)
            eh_tarde = _tarde(t, n)
            if cruza:
                intervalo_cnt[0] += 1
            if eh_tarde:
                tarde[0] += 1
            ocupado_a.add((w, d)); ocupado_b.add((w, d))
            por_semana_a.setdefault(w, []).append(disc)
            por_semana_b.setdefault(w, []).append(disc)
            resultado_a.append((w, d, t, n, disc, prof, doador_de_bloco(bl_a)))
            resultado_b.append((w, d, t, n, disc, prof, doador_de_bloco(bl_b)))
            marcas = [(tu, cessoes[tu].aplicar(w, d, t, n, disc, None))
                      for tu in (a, b)] if cessoes is not None else []
            if bt(resto):
                return True
            for (tu, marca) in marcas:
                cessoes[tu].desfazer(marca)
            resultado_a.pop(); resultado_b.pop()
            por_semana_a[w].remove(disc); por_semana_b[w].remove(disc)
            ocupado_a.discard((w, d)); ocupado_b.discard((w, d))
            if eh_tarde:
                tarde[0] -= 1
            if cruza:
                intervalo_cnt[0] -= 1
        return False

    try:
        ok = bt(list(exames))
    except TimeoutError:
        ok = False
    return ok, resultado_a, resultado_b, ocupado_a, por_semana_a, ocupado_b, por_semana_b


def resolver_par(a, b, seed, cessoes=None):
    """Resolve, na ordem de prioridade (intervalo > tarde > grupo1), todas
    as provas de professor comum entre as turmas irmas a/b, aplicadas
    simultaneamente. Retorna as alocacoes conjuntas mais o que cada turma
    ja tem ocupado, para o resolver() de cada turma completar o resto."""
    comuns = classificar_par(a, b)
    n_exames = len(comuns)
    for (max_intervalo, max_tarde, max_g1) in escada(n_exames, cessoes):
        tentativa = {tu: cessoes[tu].clone() for tu in (a, b)} \
            if cessoes is not None else None
        (ok, res_a, res_b, ocup_a, psem_a, ocup_b, psem_b) = _tentar_par(
            a, b, seed + 17 * max_tarde + 131 * max_intervalo,
            max_g1, max_tarde, max_intervalo, tentativa)
        if ok:
            if cessoes is not None:
                for tu in (a, b):
                    cessoes[tu].adotar(tentativa[tu])
            return res_a, res_b, ocup_a, psem_a, ocup_b, psem_b, comuns
    print(f"  !! {a}/{b}: NAO foi possivel coordenar as provas de professor comum")
    return [], [], set(), {}, set(), {}, comuns


# ---------------------------------------------------------------- ESCRITA
def rotulo_tempo(t, n):
    if n == 1:
        return f"{t}º tempo"
    if n == 2:
        return f"{t}º e {t+1}º tempos"
    return f"{t}º ao {t+n-1}º tempos"


def escrever(proposta, alocacoes):
    dst = os.path.join(OUT, f"Proposta_{proposta}_Calendario_Provas_2026_2SEM.xlsx")
    shutil.copy(SRC, dst)
    wb = openpyxl.load_workbook(dst)

    base = wb["10C2"]
    for turma in GRADES:
        if turma == "10C2":
            continue
        idx = wb.sheetnames.index(turma)
        del wb[turma]
        nova = wb.copy_worksheet(base)
        nova.title = turma
        nova["A1"] = f"Prüfungsplan Klasse {turma} | Plano das Provas Turma {turma}"
        # o template e a aba 10C2: apaga o AG10 dela nas turmas de outra serie
        serie = "9" if turma.startswith("9") else turma[:2]
        for r in range(3, nova.max_row + 1):
            for c in "EFGHIJ":
                v = nova[f"{c}{r}"].value
                if not v:
                    continue
                t = str(v).strip().replace("\n", " ")
                if t.startswith(("AG9", "AG10", "S1-", "S2-", "S3-", "S4-", "EX")):
                    nova[f"{c}{r}"] = None            # simulado da outra serie
                elif t.startswith(("2CH", "CC")):
                    series = [x.strip() for x in t.split(None, 1)[-1].split(",")]
                    if serie not in series:
                        nova[f"{c}{r}"] = None
        wb.move_sheet(turma, offset=idx - wb.sheetnames.index(turma))

    for turma, itens in alocacoes.items():
        ws = wb[turma]
        for (w, d, t, n, disc, prof, doador) in itens:
            r = week_row(w)
            c = COL[d]
            cell = ws[f"{c}{r}"]
            if cell.value:      # nunca sobrescrever
                print(f"  !! {turma}: {disc} sem 'celula livre' em "
                      f"sem {w} {DIAS[d]} (ja ocupada)")
                continue
            cell.value = f"{NOME[disc]} - {prof}\n\n{rotulo_tempo(t, n)}"
            cell.alignment = Alignment(wrap_text=True, vertical="center",
                                       horizontal="center")
            if cruza_intervalo(t, n):
                cell.fill = DESTAQUE_INTERVALO
        for (w, d, cod, lab) in SIMULADOS[turma]:
            r = week_row(w)
            c = COL[d]
            cell = ws[f"{c}{r}"]
            atual = str(cell.value or "").strip()
            if atual and not atual.startswith(("AG9", "AG10", "S1-", "S2-",
                                               "S3-", "S4-", "EX")):
                continue                       # celula com outra informacao
            cell.value = f"{cod}\n\n{lab}"
            cell.alignment = Alignment(wrap_text=True, vertical="center",
                                       horizontal="center")
    wb.save(dst)
    return dst


def relatorio(alocacoes_por_proposta, comuns_por_par):
    irma = {a: b for a, b in PARES_IRMAS}
    irma.update({b: a for a, b in PARES_IRMAS})

    linhas = ["# Relatório de trocas de tempo entre professores",
              "",
              "Provas de tempos seguidos em que algum tempo pertence a outra",
              "disciplina. O professor doador precisa ceder o tempo naquele dia.",
              ""]

    linhas += ["## Provas coordenadas entre turmas irmãs", "",
               "Disciplinas com o mesmo professor nas duas turmas de uma série",
               "(ver regra na skill). \"Já combinada\" é grupo paralelo que já",
               "abrange as duas turmas hoje (Alemão/DaF, GL...); \"coordenada\"",
               "precisou de um dia/tempo em comum novo, porque o professor dá",
               "aula em horários diferentes em cada turma.", "",
               "| Turmas | Disciplina | Situação |",
               "|---|---|---|"]
    for (a, b), comuns in comuns_por_par.items():
        for disc, (tipo, _profs) in sorted(comuns.items()):
            situ = "Já combinada (grupo paralelo)" if tipo == "combinada" \
                else "Coordenada (professor comum, tempos diferentes)"
            linhas.append(f"| {a}/{b} | {NOME.get(disc, disc)} | {situ} |")
    linhas.append("")

    for prop, alocacoes in alocacoes_por_proposta.items():
        linhas += [f"## Proposta {prop}", "",
                   "| Turma | Disciplina/Prof. solicitante | Tempo necessário | "
                   "Prof. doador | Disciplina do tempo doado | Ação | Observação |",
                   "|---|---|---|---|---|---|---|"]
        for turma, itens in alocacoes.items():
            comuns = comuns_por_par.get(tuple(sorted((turma, irma.get(turma, turma)))), {})
            for (w, d, t, n, disc, prof, doador) in sorted(itens):
                obs = []
                if cruza_intervalo(t, n):
                    obs.append("⚠ Cruza o intervalo do recreio — nenhuma "
                                "outra combinação coube")
                if disc in comuns:
                    tipo, _ = comuns[disc]
                    if tipo == "coordenar":
                        obs.append(f"Prova conjunta com {irma[turma]} "
                                   "(professor comum, tempos coordenados)")
                    else:
                        obs.append(f"Grupo paralelo já combinado com {irma[turma]}")
                obs_txt = "; ".join(obs)
                if not doador:
                    if obs_txt:
                        linhas.append(f"| {turma} | {NOME[disc]} / {prof} | — | — | — | — | {obs_txt} |")
                    continue
                doadores = doador if isinstance(doador[0], tuple) else (doador,)
                for (dd, dp) in doadores:
                    tempos = [t + k for k in range(n)
                              if GRADES[turma].get((d, t + k), (None,))[0] == dd]
                    tl = ", ".join(f"{x}º" for x in tempos) or f"{t+1}º"
                    linhas.append(
                        f"| {turma} | {NOME[disc]} / {prof} | {tl} tempo(s) "
                        f"({DIAS[d]}, semana {w}) | {dp} | {NOME.get(dd, dd)} | "
                        f"Solicitar ao prof. {dp} a cessão do(s) tempo(s) {tl} de "
                        f"{NOME.get(dd, dd)} para a prova de {NOME[disc]} | {obs_txt} |")
        linhas.append("")
    p = os.path.join(OUT, "Relatorio_trocas_de_tempo.md")
    with open(p, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))
    return p


def montar_proposta(seed, folga=None, sem_regra3=(), sem_regra4=()):
    """Resolve as 8 turmas. folga=None: sem os limites de cessao da
    Proposta 3. folga=N: com os limites, afrouxados em N unidades.

    sem_regra3/sem_regra4: turmas em que essas regras ficam relaxadas. O
    afrouxamento e POR TURMA -- se so uma turma nao fecha, as outras sete
    continuam cumprindo a regra integralmente.

    Devolve (alocacoes, turmas_que_nao_fecharam).
    """
    aloc = {}
    preocup_dia, pre_psem, excluir_por_turma = {}, {}, {}
    cessoes = {t: Cessoes(t, folga, regra3=t not in sem_regra3,
                          regra4=t not in sem_regra4)
               for t in GRADES} if folga is not None else None

    # 1) provas de professor comum entre turmas irmas, coordenadas ou
    #    ja combinadas -- decididas antes, para as duas turmas de vez.
    for (a, b) in PARES_IRMAS:
        res_a, res_b, ocup_a, psem_a, ocup_b, psem_b, comuns = resolver_par(
            a, b, seed + sum(map(ord, a + b)), cessoes)
        aloc[a] = list(res_a)
        aloc[b] = list(res_b)
        preocup_dia[a], pre_psem[a] = ocup_a, psem_a
        preocup_dia[b], pre_psem[b] = ocup_b, psem_b
        excluir_por_turma[a] = set(comuns)
        excluir_por_turma[b] = set(comuns)
        if len(res_a) < len(comuns):
            print(f"  !! {a}/{b}: só {len(res_a)} de {len(comuns)} provas "
                  "comuns foram coordenadas")

    # 2) o resto de cada turma (disciplinas independentes, ex. Ingles
    #    nas turmas 10/11/12), preenchendo ao redor do que ja foi
    #    decidido acima.
    falharam = set()
    for turma in GRADES:
        ok, res, g1, tarde, cruzadas = resolver(
            turma, seed + sum(map(ord, turma)),
            preocupado_dia=preocup_dia.get(turma),
            pre_por_semana=pre_psem.get(turma),
            excluir=excluir_por_turma.get(turma),
            cessoes=cessoes[turma] if cessoes else None)
        if not ok:
            falharam.add(turma)
        else:
            avisos = []
            if g1 > 1:
                avisos.append(f"{g1} disciplinas do grupo 1 na mesma semana")
            if tarde:
                tardias = ", ".join(
                    f"{NOME[dc]} ({rotulo_tempo(t, n)})"
                    for (_w, _d, t, n, dc, *_r) in sorted(res) if _tarde(t, n))
                avisos.append(f"{tarde} prova(s) nos tempos 7-11: {tardias}")
            if cruzadas:
                avisos.append(f"{cruzadas} prova(s) cruzando o intervalo do recreio")
            if avisos:
                print(f"  ~  {turma}: " + "; ".join(avisos))
        aloc[turma].extend(res)
    return aloc, falharam


def main():
    os.makedirs(OUT, exist_ok=True)
    carregar_ocupadas()

    comuns_por_par = {(a, b): classificar_par(a, b) for a, b in PARES_IRMAS}

    todas = {}
    for prop, seed in [(1, 20260806), (2, 99887766)]:
        aloc, _ok = montar_proposta(seed)
        caminho = escrever(prop, aloc)
        todas[prop] = aloc
        print(f"Proposta {prop} -> {caminho}")

    # Proposta 3: mesmas regras das anteriores MAIS os limites de cessao de
    # aula (regras 1 a 5 pedidas pela coordenacao). Tenta primeiro com os
    # limites estritos; se nao fechar, afrouxa uma unidade por vez, para
    # entregar a solucao mais proxima possivel do pedido.
    # Escada de afrouxamento: primeiro sobe o teto de cessoes; so depois
    # relaxa a regra 4 (nao ceder as vesperas da propria prova) e, por
    # ultimo, a 3 (duas semanas sem contato). As regras 1, 2 e 5 (tetos) e
    # as datas exigidas pela coordenacao nunca sao relaxadas.
    print("Proposta 3: aplicando os limites de cessão de aula")
    aloc3, sem_r3, sem_r4, folga = None, set(), set(), 0
    for etapa in range(0, 12):
        aloc, falharam = montar_proposta(20261107, folga=folga,
                                         sem_regra3=sem_r3, sem_regra4=sem_r4)
        if aloc3 is None:
            aloc3 = aloc                      # guarda a 1a como reserva
        if not falharam:
            aloc3 = aloc
            break
        aloc3 = aloc
        # Afrouxa na ordem inversa da importancia. Os tetos (regras 1, 2 e
        # 5) sao limites duros dados em numero pela escola, entao a folga
        # neles e o ULTIMO recurso: subir a folga vale para todas as turmas
        # e estoura varios tetos de uma vez, enquanto relaxar a regra 4
        # atinge so a turma que nao fecha.
        if not sem_r4 >= falharam:
            sem_r4 |= falharam
            print(f"  ~  relaxando a regra 4 apenas em {sorted(falharam)}")
        elif not sem_r3 >= falharam:
            sem_r3 |= falharam
            print(f"  ~  relaxando a regra 3 apenas em {sorted(falharam)}")
        elif folga < 3:
            folga += 1
            print(f"  ~  {sorted(falharam)} ainda não fecharam; subindo o "
                  f"teto de cessões para +{folga}")
        else:
            print(f"  !! {sorted(falharam)} não fecharam nem com tudo relaxado")
            break
    if folga or sem_r3 or sem_r4:
        print(f"  ATENÇÃO: Proposta 3 fechada com folga +{folga}"
              + (f"; regra 4 relaxada em {sorted(sem_r4)}" if sem_r4 else "")
              + (f"; regra 3 relaxada em {sorted(sem_r3)}" if sem_r3 else "")
              + ". As demais turmas cumprem todas as regras.")
    else:
        print("  Proposta 3 fechada com todos os limites estritos")
    caminho = escrever(3, aloc3)
    todas[3] = aloc3
    print(f"Proposta 3 -> {caminho}")

    print("Relatorio ->", relatorio(todas, comuns_por_par))


if __name__ == "__main__":
    main()
