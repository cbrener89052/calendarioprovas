#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Cruza as provas do 2o semestre de 2025 (analisar_2sem_2025.py) com a
grade de aulas daquele semestre (limpar_grade_2025.py) e conta as
cessoes, no mesmo formato dos relatorios de 2026.

O calendario de 2025 registra o DIA da prova, mas nao os tempos -- ao
contrario de 2026 e do 1o semestre, aqui os tempos sao RECONSTRUIDOS,
pelas regras que a escola confirmou:

  1. Duracao = o numero de tempos que a propria celula do calendario ja
     traz entre parenteses (dado do arquivo, nao suposicao).
  2. Tempos proprios = as aulas que a disciplina tem com aquela turma,
     naquele dia da semana (todas contam, nao so as necessarias).
  3. Tempo emprestado = o que faltar para completar a duracao vem do
     tempo VIZINHO ao tempo proprio mais perto (antes ou depois),
     evitando doador de aula unica na semana.
  4. Provas no mesmo dia entre turmas irmas sao a MESMA prova aplicada
     ao mesmo tempo -- entao quando uma turma nao tem nenhum tempo
     proprio da disciplina naquele dia mas a turma irma tem, a ancora
     vem da turma irma (a prova aconteceu no horario natural dela).

Import ANTES de rodar: python extrair_grade_2025.py --tsv (uma vez) e
depois python limpar_grade_2025.py.
"""
import os, sys, collections
import analisar_2sem_2025 as P

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "Horario desenvolvido")
sys.path.insert(0, os.path.join(BASE, "horarios2025"))
from grade_2sem_2025_limpa import GRADE_2025

IRMA = {"9C1": "9C2", "9C2": "9C1", "10C1": "10C2", "10C2": "10C1",
        "11C1": "11C2", "11C2": "11C1"}          # 12C1 nao tem irma em 2025

# Disciplina examinada -> codigos da grade que sao "tempo proprio" dela.
FAMILIA = {"ing": {"ing", "ingT"}}


def familia_de(disc):
    return FAMILIA.get(disc, {disc})


def tempos_proprios(turma, dia, disc):
    fam = familia_de(disc)
    g = GRADE_2025.get(turma, {})
    out = []
    for (d, t), pares in g.items():
        if d == dia and any(c in fam for (c, _p) in pares):
            out.append(t)
    return sorted(out)


def nao_doadoras(turma):
    """Codigos que so tem 1 aula na semana naquela turma -- nao podem
    ceder (perderiam a unica aula)."""
    cont = collections.Counter()
    for pares in GRADE_2025.get(turma, {}).values():
        for (c, _p) in pares:
            cont[c] += 1
    return {c for c, n in cont.items() if n == 1}


def dono_do_tempo(turma, dia, tempo):
    """(codigo, professor) mais provavel daquele tempo -- so o primeiro
    professor quando ha grupos paralelos (varios profs, mesmo tempo);
    para cessao o que importa e a disciplina doadora, nao qual professor
    do grupo especificamente."""
    pares = GRADE_2025.get(turma, {}).get((dia, tempo))
    return pares[0] if pares else None


def reconstruir(turma, dia, disc, n_tempos):
    """(tempos_usados, cedidas, pendencia). cedidas: [((codigo,prof), t)].
    pendencia != None quando nao ha ancora nenhuma (nem na turma, nem na
    irma) para reconstruir com confianca."""
    proprios = tempos_proprios(turma, dia, disc)
    ancora_turma = turma
    if not proprios:
        irma = IRMA.get(turma)
        if irma:
            proprios_irma = tempos_proprios(irma, dia, disc)
            if proprios_irma:
                proprios = proprios_irma
                ancora_turma = irma      # so para saber de onde veio a ancora
    if not proprios:
        return None, None, "sem tempo próprio nesta turma nem na irmã"

    usados = list(proprios[:n_tempos])
    veto = nao_doadoras(turma)
    cedidas = []
    tentativas_sem_sucesso = []
    while len(usados) < n_tempos:
        base = usados[-1] if usados[-1] > proprios[0] else usados[0]
        candidatos = []
        for ref in (usados[0], usados[-1]):
            candidatos += [ref - 1, ref + 1]
        candidatos = [t for t in candidatos
                      if 1 <= t <= 11 and t not in usados]
        escolhido = None
        for t in sorted(set(candidatos)):
            dono = dono_do_tempo(turma, dia, t)
            if dono is None:
                continue
            if dono[0] in familia_de(disc):
                continue
            if dono[0] in veto:
                tentativas_sem_sucesso.append((t, dono, "não doa (1x/semana)"))
                continue
            escolhido = (t, dono)
            break
        if escolhido is None:
            return usados, cedidas, (
                f"faltam {n_tempos - len(usados)} tempo(s) e nenhum vizinho "
                f"válido foi encontrado")
        t, dono = escolhido
        usados.append(t)
        cedidas.append((dono, t))
    return sorted(usados), cedidas, None


def escrever(tabelas):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from exportar_tabelas_turma import carregar_siglas, nomes_dos_profs
    siglas = carregar_siglas()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cab_fill = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    zebra = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="D9E1F2")

    for turma in P.TURMAS:
        ws = wb.create_sheet(turma)
        ws["A1"] = (f"Tempos Cedidos para Provas — {turma} — "
                    f"2º semestre 2025 (o que de fato ocorreu; tempos das "
                    f"provas reconstruídos, ver observação)")
        ws["A1"].font = Font(bold=True, size=12)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)

        for c, t in enumerate(["Disciplina", "Professor(es)",
                               "Nº de aulas semanais",
                               "Nº de aulas programadas no semestre",
                               "Nº de aulas cedidas para provas de outras disciplinas",
                               "% de aulas cedidas no semestre"], 1):
            cel = ws.cell(2, c, t)
            cel.fill, cel.font = cab_fill, Font(bold=True, color="FFFFFF", size=11)
            cel.alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=True)
            cel.border = borda

        i = 3
        for (disc, prof, n_sem, n_prog, n_ced, pct) in tabelas[turma]:
            ws.cell(i, 1, P.NOME_EXIBIDO.get(disc, disc))
            ws.cell(i, 2, "—" if prof in ("-", "?") else
                    nomes_dos_profs(prof, siglas))
            ws.cell(i, 3, n_sem)
            ws.cell(i, 4, n_prog)
            ws.cell(i, 5, n_ced)
            cel = ws.cell(i, 6, pct if pct is not None else "—")
            if pct is not None:
                cel.number_format = "0.0%"
            for c in range(1, 7):
                x = ws.cell(i, c)
                x.border = borda
                x.alignment = Alignment(vertical="center", wrap_text=(c == 2),
                                        horizontal="center" if c != 2 else "left")
                if i % 2:
                    x.fill = zebra
            i += 1

        tp = sum(l[3] for l in tabelas[turma])
        tc = sum(l[4] for l in tabelas[turma])
        ws.cell(i, 1, "Total").font = Font(bold=True)
        ws.cell(i, 4, tp)
        ws.cell(i, 5, tc)
        cel = ws.cell(i, 6, (tc / tp) if tp else "—")
        if tp:
            cel.number_format = "0.0%"
        for c in range(1, 7):
            x = ws.cell(i, c)
            x.border, x.fill, x.font = borda, total_fill, Font(bold=True)
            x.alignment = Alignment(horizontal="center" if c != 2 else "left")

        for col, larg in zip("ABCDEF", (16, 46, 14, 18, 16, 14)):
            ws.column_dimensions[col].width = larg
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:F{i}"

    destino = os.path.join(OUT, "Relatorio_Tempos_Cedidos_2SEM_2025.xlsx")
    wb.save(destino)
    return destino


def main():
    provas, ilegiveis = P.ler_provas()
    fora = [p for p in provas if not P.dentro_do_periodo(p[0], p[1])]
    provas = [p for p in provas if P.dentro_do_periodo(p[0], p[1])]
    livres = P.dias_sem_aula()

    cedidas = collections.defaultdict(collections.Counter)
    pendencias = []
    for (turma, data, dia, disc, n, txt) in provas:
        usados, ced, pend = reconstruir(turma, dia, disc, n)
        if pend:
            pendencias.append((turma, data, disc, n, txt, pend))
            continue
        for (dono, _t) in ced:
            cedidas[turma][dono] += 1

    print(f"Provas dentro do período letivo: {len(provas)} "
          f"(+{len(fora)} fora do período, tratadas como sem cessão — "
          "aulas regulares já haviam terminado)")
    print(f"{len(pendencias)} prova(s) sem reconstrução possível:")
    for (t, dt, disc, n, txt, pend) in pendencias:
        print(f"   {t} {dt} {disc}({n}) — {pend} — {txt!r}")

    tabelas = {}
    for turma in P.TURMAS:
        semanas = P.semanas_letivas(turma, livres)
        semanal = collections.Counter()
        for pares in GRADE_2025.get(turma, {}).values():
            for (c, prof) in pares:
                semanal[(c, prof)] += 1
        programadas = collections.Counter()
        for (d, _t), pares in GRADE_2025.get(turma, {}).items():
            for (c, prof) in pares:
                programadas[(c, prof)] += semanas.get(d, 0)

        linhas = []
        for chave, n_sem in semanal.items():
            n_prog = programadas[chave]
            n_ced = cedidas[turma].get(chave, 0)
            linhas.append((chave[0], chave[1], n_sem, n_prog, n_ced,
                           (n_ced / n_prog) if n_prog else None))
        linhas.sort(key=lambda x: (-x[4], x[0], x[1]))
        tabelas[turma] = linhas

    destino = escrever(tabelas)
    print("\nGravado:", destino)
    for turma in P.TURMAS:
        tp = sum(l[3] for l in tabelas[turma])
        tc = sum(l[4] for l in tabelas[turma])
        pior = tabelas[turma][0] if tabelas[turma] else None
        linha = f"  {turma}: {tc} de {tp} aulas cedidas ({tc / tp:.1%})" if tp else f"  {turma}: sem dados"
        if pior and pior[4]:
            linha += f" | pior: {pior[0]}/{pior[1]} {pior[4]}x = {pior[5]:.1%}"
        print(linha)


if __name__ == "__main__":
    main()
