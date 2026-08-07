#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Conta as aulas cedidas para provas no 2o semestre de 2025.

E a comparacao mais justa para o 2o semestre de 2026 planejado: mesmo
semestre do calendario escolar, mesma dinamica de viagem das turmas 10 e
12. Ver referencia/analise_2sem_2025.md.

O calendario de 2025 e mais limpo que o de 2026: uma linha por semana e a
celula ja traz o NUMERO DE TEMPOS entre parenteses -- 'FIS (2)',
'PORT (3)', 'Soc (1)'. O que ele NAO traz e QUAIS tempos, e por isso a
posicao da prova na grade precisa ser reconstruida (ver localizar()).

Em 2025 havia 7 turmas: nao existia 12C2.
"""
import openpyxl, os, re, collections, datetime, unicodedata

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "provas2sem_2025",
                   "Klausurplan_ramoC_2025_2SEM.xlsx")
OUT = os.path.join(BASE, "Horario desenvolvido")

TURMAS = ["9C1", "9C2", "10C1", "10C2", "11C1", "11C2", "12C1"]

# As abas nao tem o mesmo layout: a 12C1 esta uma coluna a esquerda das
# demais. Por isso as colunas dos dias sao localizadas pelo cabecalho, e
# a data e procurada como a primeira celula de data da linha.
DIAS_CAB = {"montag": 1, "dienstag": 2, "mittwoch": 3, "donnerstag": 4,
            "freitag": 5}

# Periodo letivo confirmado pela escola. As turmas 10 e 12 viajam antes.
INICIO = datetime.date(2025, 7, 28)
FIM = {"9_11": datetime.date(2025, 11, 28),
       "10_12": datetime.date(2025, 11, 7)}

# Marcacoes que nao sao prova de disciplina: simulados, avaliacoes
# externas, conselho de classe, avisos administrativos.
NAO_PROVA = (
    "unterrichtsfrei", "2 cha", "2cha", "cha", "ag", "ar", "aulao", "coe",
    "carta", "dsd", "fer", "ic", "ielts", "in10", "rgp", "s3-", "s4-",
    "sat", "vest", "viagem", "cc", "prazo", "segunda chamada", "f12",
)

# Como a disciplina aparece -> codigo. A ordem importa: 'pgram' e 'pred'
# antes de 'port', senao o prefixo curto casaria primeiro.
ALIAS = [
    ("pgram", "gram"), ("pred", "pred"), ("port", "port"),
    ("mat", "mat"), ("qui", "qui"), ("fis", "fis"), ("bio", "bio"),
    ("geo", "geo"), ("hist", "his"), ("his", "his"), ("ing", "ing"),
    ("daf", "DaF"), ("gl", "GL"), ("fil", "fil"), ("soc", "soc"),
]


def sem_acento(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def grupo(turma):
    return "9_11" if turma.startswith(("9", "11")) else "10_12"


def colunas_da_aba(ws):
    """(linha do cabecalho, {coluna: dia}) — varia de aba para aba."""
    for r in range(1, 6):
        col = {}
        for c in range(1, ws.max_column + 1):
            v = sem_acento(str(ws.cell(r, c).value or "").strip().lower())
            if v in DIAS_CAB:
                col[c] = DIAS_CAB[v]
        if len(col) == 5:
            return r, col
    raise SystemExit("cabeçalho de dias não encontrado")


def ler_provas(caminho=SRC):
    """[(turma, data, dia, disciplina, n_tempos, texto)] + nao lidas."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    provas, ilegiveis = [], []
    for turma in TURMAS:
        ws = wb[turma]
        r_cab, cols = colunas_da_aba(ws)
        primeira_col_dia = min(cols)
        for r in range(r_cab + 1, ws.max_row + 1):
            # a data de inicio da semana e a primeira data da linha, a
            # esquerda das colunas de dia
            segunda = None
            for c in range(1, primeira_col_dia):
                v = ws.cell(r, c).value
                if isinstance(v, datetime.datetime):
                    segunda = v.date()
                    break
            if segunda is None:
                continue
            for c, d in cols.items():
                v = ws.cell(r, c).value
                if not v or not str(v).strip():
                    continue
                txt = " ".join(str(v).split())
                baixo = sem_acento(txt.lower()).lstrip("*• ").strip()
                if any(re.search(rf"(?<![a-z]){re.escape(x)}", baixo)
                       for x in NAO_PROVA):
                    continue
                m = re.search(r"\((\d)\)", txt)
                disc = next((cod for alias, cod in ALIAS
                             if re.match(rf"{alias}", baixo)), None)
                if not m or not disc:
                    ilegiveis.append((turma, segunda, d, txt))
                    continue
                data = segunda + datetime.timedelta(days=d - 1)
                provas.append((turma, data, d, disc, int(m.group(1)), txt))
    return provas, ilegiveis


def dentro_do_periodo(turma, data):
    return INICIO <= data <= FIM[grupo(turma)]


def main():
    provas, ilegiveis = ler_provas()
    fora = [p for p in provas if not dentro_do_periodo(p[0], p[1])]
    print(f"Provas lidas: {len(provas)}")
    print(f"Fora do período letivo: {len(fora)}")
    for p in fora:
        print(f"   ! {p[0]} {p[1]} {p[3]} — depois do fim das aulas")
    print(f"Não interpretadas: {len(ilegiveis)}")
    for x in ilegiveis:
        print("   ?", x)
    for t in TURMAS:
        cont = collections.Counter(
            f"{p[3]}({p[4]})" for p in provas if p[0] == t)
        print(f"  {t}: {sum(cont.values())} provas — " +
              ", ".join(f"{k}:{v}" for k, v in sorted(cont.items())))


if __name__ == "__main__":
    main()
