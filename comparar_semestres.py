#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compara as cessoes de aula do 1o semestre (o que ocorreu) com as das
propostas do 2o semestre.

Cuidado ao ler os numeros ABSOLUTOS: os semestres tem tamanhos diferentes
(23 semanas letivas no 1o; 15 nas turmas 10/12 e 17 nas 9/11 no 2o) e as
grades tambem mudaram. Por isso a comparacao honesta e o PERCENTUAL de
aulas cedidas, e nao a contagem -- um mesmo numero de provas se espalha
por menos aulas no 2o semestre, o que sozinho ja empurra o percentual
para cima.

Grava: Horario desenvolvido/Comparativo_Cessoes_1SEM_x_Proposta3.xlsx
"""
import openpyxl, os, collections
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "Horario desenvolvido")
TURMAS = ["9C1", "9C2", "10C1", "10C2", "11C1", "11C2", "12C1", "12C2"]


def ler(caminho):
    """{(turma, disciplina, sigla): (sem, prog, cedidas, pct)}"""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    out = {}
    for t in wb.sheetnames:
        for row in wb[t].iter_rows(min_row=3, values_only=True):
            if row[0] is None or row[0] == "Total":
                continue
            sigla = str(row[1]).split(" (")[0].strip() if row[1] else "—"
            out[(t, row[0], sigla)] = (row[2], row[3], row[4], row[5])
    return out


def main():
    s1 = ler(os.path.join(OUT, "Relatorio_Tempos_Cedidos_1SEM.xlsx"))
    p3 = ler(os.path.join(OUT, "Relatorio_Tempos_Cedidos_Proposta_3.xlsx"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    ws["A1"] = ("Aulas cedidas: 1º semestre (ocorrido) x Proposta 3 "
                "(2º semestre)")
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = ("Compare pelo percentual: o 1º semestre teve 23 semanas "
                "letivas e o 2º tem 15 (turmas 10/12) ou 17 (turmas 9/11), "
                "então o mesmo número de provas pesa mais no 2º.")
    ws["A2"].font = Font(italic=True, size=10)
    ws.merge_cells("A2:H2")
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    cab = ["Turma", "Disciplina", "Professor",
           "1º sem: cedidas", "1º sem: %", "Proposta 3: cedidas",
           "Proposta 3: %", "Variação (p.p.)"]
    cab_fill = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for c, t in enumerate(cab, 1):
        cel = ws.cell(3, c, t)
        cel.fill = cab_fill
        cel.font = Font(bold=True, color="FFFFFF", size=11)
        cel.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        cel.border = borda

    pior = PatternFill("solid", fgColor="F8CBAD")   # piorou
    melhor = PatternFill("solid", fgColor="C6E0B4")  # melhorou

    linhas = []
    for chave in set(s1) | set(p3):
        a, b = s1.get(chave), p3.get(chave)
        if not a or not b:
            continue                 # disciplina/professor só existe num semestre
        if not a[2] and not b[2]:
            continue                 # não cedeu em nenhum dos dois
        linhas.append((chave, a, b, (b[3] or 0) - (a[3] or 0)))
    linhas.sort(key=lambda x: -abs(x[3]))

    i = 4
    for ((turma, disc, sigla), a, b, delta) in linhas:
        ws.cell(i, 1, turma)
        ws.cell(i, 2, disc)
        ws.cell(i, 3, sigla)
        ws.cell(i, 4, a[2])
        ws.cell(i, 5, a[3]).number_format = "0.0%"
        ws.cell(i, 6, b[2])
        ws.cell(i, 7, b[3]).number_format = "0.0%"
        cel = ws.cell(i, 8, delta)
        cel.number_format = "+0.0%;-0.0%"
        if abs(delta) > 0.005:
            cel.fill = pior if delta > 0 else melhor
        for c in range(1, 9):
            x = ws.cell(i, c)
            x.border = borda
            x.alignment = Alignment(horizontal="center" if c != 3 else "left")
        i += 1

    for col, larg in zip("ABCDEFGH", (10, 14, 12, 15, 12, 17, 14, 15)):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{i - 1}"

    destino = os.path.join(OUT, "Comparativo_Cessoes_1SEM_x_Proposta3.xlsx")
    wb.save(destino)
    print("Gravado:", destino)
    print(f"{len(linhas)} pares (turma, disciplina, professor) comparáveis")
    piores = sorted(linhas, key=lambda x: -x[3])[:6]
    print("\nMaiores aumentos de percentual:")
    for ((t, d, s), a, b, dl) in piores:
        print(f"  {t} {d}/{s}: {a[3]:.1%} -> {b[3]:.1%}  ({dl:+.1%})")
    print("\nMaiores reduções:")
    for ((t, d, s), a, b, dl) in sorted(linhas, key=lambda x: x[3])[:6]:
        print(f"  {t} {d}/{s}: {a[3]:.1%} -> {b[3]:.1%}  ({dl:+.1%})")


if __name__ == "__main__":
    main()
