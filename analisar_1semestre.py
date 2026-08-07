#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Conta as aulas cedidas para provas no 1o semestre de 2026 (o que de fato
ocorreu), no mesmo formato do relatorio do 2o semestre.

Diferente do 2o semestre, aqui nada e planejado: o calendario de provas ja
foi executado e esta em Horario modelo/Klausurplan_2026_1SEM.xlsx. O
trabalho e ler o que aconteceu e cruzar com a grade de aulas do 1o
semestre (que e DIFERENTE da do 2o).

Saida, uma aba por turma:
  Disciplina | Professor | aulas/semana | aulas no semestre | cedidas | %

Cuidado com o arquivo do 1o semestre: o preenchimento foi manual e e
bem irregular -- os tempos aparecem ora na linha do titulo, ora na da
sala, com ou sem "º", por extenso ("primeiro e segundo"), como fracao
("4.5") ou como intervalo ("1-7"). Por isso o parser abaixo e tolerante,
e tudo que ele NAO conseguir ler com seguranca e reportado em vez de
adivinhado.
"""
import openpyxl, os, re, collections, datetime, unicodedata

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "Horario modelo", "Klausurplan_2026_1SEM.xlsx")
OUT = os.path.join(BASE, "Horario desenvolvido")

TURMAS = ["9C1", "9C2", "10C1", "10C2", "11C1", "11C2", "12C1", "12C2"]
COLS = {"E": 1, "F": 2, "G": 3, "H": 4, "I": 5}
SEMANA1 = datetime.date(2026, 2, 2)     # segunda da semana 1 do 1o semestre

# Marcacoes que NAO sao prova de disciplina (nao consomem tempo de colega
# de forma atribuivel a uma disciplina).
NAO_PROVA = ("unterrichtsfrei", "fach |", "matéria |", "materia |", "cc ",
             "2ch", "viagem", "sim ", "ag1", "ag9", "ag10", "s1-", "s2-",
             "s3-", "s4-", "sip", "conselho", "recesso", "feriado",
             "alemun", "dsd")

EXTENSO = {"primeiro": 1, "segundo": 2, "terceiro": 3, "quarto": 4,
           "quinto": 5, "sexto": 6, "setimo": 7, "oitavo": 8, "nono": 9,
           "decimo": 10}


def week_row(w):
    return 3 + 3 * (w - 1)


def sem_acento(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def _limpar(txt):
    """Normaliza o texto para leitura dos tempos.

    Tira acentos, remove codigos de sala (E301, A 314...) para os numeros
    deles nao virarem 'tempos', e uniformiza o ordinal: no arquivo o mesmo
    tempo aparece como '1º', '1o', '1°' e '1ª'.
    """
    t = sem_acento(txt.lower())
    t = re.sub(r"\b[a-z]\s*\d{3}\b", " ", t)        # sala: E301, A 304
    t = re.sub(r"(\d)\s*[o°ºa](?![a-z0-9])", r"\1º", t)
    return t


def tempos_do_texto(txt):
    """Lista de tempos (1..11) que a prova ocupou, lida de um texto solto.

    Devolve [] quando nao houver nada reconhecivel -- nunca chuta.
    """
    # so os trechos que falam de tempo/aula interessam; 'Sala de aula' e
    # 'Sala de provas' sao descricao de local e nao podem entrar
    partes = []
    for seg in txt.split("|"):
        s = _limpar(seg)
        if re.match(r"\s*sala\b", s) and not re.search(r"temp", s):
            continue
        partes.append(s)
    t = " ".join(partes)

    # "1-7", "1 a 7": intervalo
    m = re.search(r"\b(\d{1,2})\s*º?\s*(?:-|a|ao|ate)\s*(\d{1,2})\s*º?\s*temp", t)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 1 <= a <= b <= 11:
            return list(range(a, b + 1))

    # por extenso: "primeiro e segundo tempos"
    achados = [v for k, v in EXTENSO.items() if re.search(rf"\b{k}\b", t)]
    if achados:
        return sorted(set(achados))

    # "4.5" ou "2,3" (duas aulas coladas por ponto/virgula)
    m = re.search(r"(?<!\d)(\d)\s*[.,]\s*(\d)(?!\d)", t)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 1 <= a <= 11 and 1 <= b <= 11:
            return sorted({a, b})

    # forma geral: numeros num trecho que fala de tempo/aula --
    # "4º, 5 e 6º tempos", "2 e 3 tempos", "6 tempo", "4º e 5º temp"
    if re.search(r"temp|aula", t):
        nums = [int(x) for x in re.findall(r"(?<!\d)(\d{1,2})(?!\d)", t)]
        nums = [n for n in nums if 1 <= n <= 11]
        if nums:
            return sorted(set(nums))

    # numeros com marcador de ordinal, mesmo sem a palavra "tempo"
    nums = [int(x) for x in re.findall(r"(?<!\d)(\d{1,2})\s*º", t)]
    nums = [n for n in nums if 1 <= n <= 11]
    return sorted(set(nums))


# Como a disciplina aparece escrita no arquivo -> codigo usado na grade.
# No 1o semestre o preenchimento foi livre, entao a mesma disciplina
# aparece de varias formas ("MAT", "MATEMÁTICA", "MAT/ClaMe e JJ").
ALIAS = [
    ("lp/lit/red", "LPLITRED"), ("pred", "pred"), ("redacao", "pred"),
    ("port", "port"), ("portugues", "port"), ("gram", "gram"),
    ("matematica", "mat"), ("mat", "mat"),
    ("quimica", "qui"), ("qui", "qui"),
    ("fisica", "fis"), ("fis", "fis"),
    ("biologia", "bio"), ("bio", "bio"),
    ("geografia", "geo"), ("geo", "geo"),
    ("historia", "his"), ("hist", "his"), ("his", "his"),
    ("ingles", "ing"), ("ing", "ing"),
    ("alemao", "DaF"), ("daf", "DaF"),
    ("gl", "GL"),
    ("filosofia", "fil"), ("fil", "fil"),
    ("sociologia", "soc"), ("soc", "soc"),
]


def disciplina_do_texto(txt):
    """Codigo da disciplina, procurado em QUALQUER linha da celula.

    No arquivo do 1o semestre a ordem das 3 linhas nao e confiavel: em
    varias celulas o tempo esta na linha do titulo e o nome da disciplina
    na de baixo (ou falta). Devolve None quando nao houver nome.
    """
    t = sem_acento(txt.lower())
    t = re.sub(r"\b[a-z]\s*\d{3}\b", " ", t)     # tira codigo de sala
    for alias, cod in ALIAS:
        if re.search(rf"(?<![a-z]){re.escape(alias)}(?![a-z])", t):
            return cod
    return None


def ler_provas(caminho=SRC):
    """[(turma, semana, dia, texto, [tempos])] do que foi aplicado.

    Tambem devolve a lista do que nao deu para interpretar.
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    provas, ilegiveis = [], []
    for turma in TURMAS:
        ws = wb[turma]
        for w in range(1, 25):
            r = week_row(w)
            if r + 2 > ws.max_row:
                break
            for col, d in COLS.items():
                v = ws[f"{col}{r}"].value
                if not v:
                    continue
                cabeca = str(v).strip()
                baixo = sem_acento(cabeca.lower())
                if any(x in baixo for x in NAO_PROVA):
                    continue
                linhas = [cabeca,
                          str(ws[f"{col}{r + 1}"].value or ""),
                          str(ws[f"{col}{r + 2}"].value or "")]
                junto = " | ".join(x for x in linhas if x.strip())
                tempos = tempos_do_texto(junto)
                disc = disciplina_do_texto(junto)
                if not tempos or not disc:
                    falta = []
                    if not tempos:
                        falta.append("tempos")
                    if not disc:
                        falta.append("disciplina")
                    ilegiveis.append((turma, w, d, junto, "+".join(falta)))
                    continue
                provas.append((turma, w, d, junto, tempos, disc))
    provas, ilegiveis = inferir_pela_irma(provas, ilegiveis)
    return provas, ilegiveis


IRMA = {"9C1": "9C2", "9C2": "9C1", "10C1": "10C2", "10C2": "10C1",
        "11C1": "11C2", "11C2": "11C1", "12C1": "12C2", "12C2": "12C1"}


def inferir_pela_irma(provas, ilegiveis):
    """Recupera a disciplina de celulas em que so o tempo foi preenchido.

    Vale so quando a turma irma tem prova no MESMO dia e nos MESMOS
    tempos: nesse caso e a mesma prova aplicada nas duas turmas (professor
    comum), e a disciplina e a mesma. Fora disso nao infere -- o caso fica
    na lista de pendencias.
    """
    porpar = {}
    for (t, w, d, _txt, tempos, disc) in provas:
        porpar[(t, w, d)] = (tuple(tempos), disc)

    resolvidas, restantes = [], []
    for item in ilegiveis:
        turma, w, d, txt, motivo = item
        if motivo != "disciplina":
            restantes.append(item)
            continue
        tempos = tempos_do_texto(txt)
        par = porpar.get((IRMA[turma], w, d))
        if par and par[0] == tuple(tempos):
            resolvidas.append((turma, w, d, txt + " [disciplina inferida da "
                               "turma irmã]", tempos, par[1]))
        else:
            restantes.append(item)
    return provas + resolvidas, restantes


# Disciplina da PROVA -> como ela aparece na grade de aulas. O tempo so
# conta como cessao quando o dono dele nao pertence a esta familia.
FAMILIA = {
    "port": {"p", "plit"},
    "LPLITRED": {"p", "plit", "pred", "gram"},
    "pred": {"pred"},
    "ing": {"ing", "ingT"},          # nas turmas 9 o ingles aparece como ingT
}

ULTIMA_SEMANA = 23                   # 06 a 10/07/2026, fim do 1o semestre


def dias_sem_aula(caminho=SRC):
    """{turma: {(semana, dia)}} marcados como 'unterrichtsfrei' no proprio
    calendario de provas — é a fonte da escola para feriados e recessos."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    out = {}
    for turma in TURMAS:
        ws = wb[turma]
        livres = set()
        for w in range(1, ULTIMA_SEMANA + 1):
            r = week_row(w)
            if r > ws.max_row:
                break
            for col, d in COLS.items():
                if "unterrichtsfrei" in str(ws[f"{col}{r}"].value or "").lower():
                    livres.add((w, d))
        out[turma] = livres
    return out


def aulas_programadas(turma, dia, livres):
    """Quantas vezes uma aula desse dia da semana aconteceu no semestre."""
    return sum(1 for w in range(1, ULTIMA_SEMANA + 1)
               if (w, dia) not in livres[turma])


def contar(grade, provas, livres):
    """{turma: [(disc, prof, aulas_sem, programadas, cedidas, pct)]}"""
    tabelas = {}
    for turma in TURMAS:
        g = grade[turma]
        semanais = collections.Counter(g.values())
        programadas = collections.Counter()
        for (d, _t), chave in g.items():
            programadas[chave] += aulas_programadas(turma, d, livres)

        cedidas = collections.Counter()
        for (t, w, d, _txt, tempos, disc) in provas:
            if t != turma:
                continue
            familia = FAMILIA.get(disc, {disc})
            for tempo in tempos:
                dono = g.get((d, tempo))
                if not dono or dono[0] in familia:
                    continue
                cedidas[dono] += 1

        linhas = []
        for chave, n_sem in semanais.items():
            n_prog = programadas[chave]
            n_ced = cedidas.get(chave, 0)
            linhas.append((chave[0], chave[1], n_sem, n_prog, n_ced,
                           (n_ced / n_prog) if n_prog else None))
        linhas.sort(key=lambda x: (-x[4], x[0], x[1]))
        tabelas[turma] = linhas
    return tabelas


def main():
    provas, ilegiveis = ler_provas()
    print(f"Provas lidas: {len(provas)}")
    print(f"Não interpretadas: {len(ilegiveis)}")
    for x in ilegiveis:
        print("   ?", x)
    import sys
    sys.path.insert(0, os.path.join(BASE, "horarios_1semestre"))
    from grade_1semestre import GRADE_1SEM

    livres = dias_sem_aula()
    tabelas = contar(GRADE_1SEM, provas, livres)
    destino = escrever(tabelas)
    print("\nGravado:", destino)
    for t in TURMAS:
        tot_p = sum(l[3] for l in tabelas[t])
        tot_c = sum(l[4] for l in tabelas[t])
        pior = tabelas[t][0]
        print(f"  {t}: {tot_c} de {tot_p} aulas cedidas ({tot_c / tot_p:.1%}) "
              f"| pior: {pior[0]}/{pior[1]} {pior[4]}x = "
              f"{pior[5]:.1%}" if tot_p else t)


def escrever(tabelas):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from exportar_tabelas_turma import carregar_siglas, nomes_dos_profs
    siglas = carregar_siglas()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cab_fill = PatternFill("solid", fgColor="1F3864")
    cab_font = Font(bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    zebra = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="D9E1F2")

    for turma in TURMAS:
        ws = wb.create_sheet(turma)
        ws["A1"] = (f"Tempos Cedidos para Provas — {turma} — "
                    f"1º semestre 2026 (o que de fato ocorreu)")
        ws["A1"].font = Font(bold=True, size=13)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center")

        for c, t in enumerate(["Disciplina", "Professor(es)",
                               "Nº de aulas semanais",
                               "Nº de aulas programadas no semestre",
                               "Nº de aulas cedidas para provas de outras disciplinas",
                               "% de aulas cedidas no semestre"], 1):
            cel = ws.cell(2, c, t)
            cel.fill, cel.font = cab_fill, cab_font
            cel.alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=True)
            cel.border = borda

        i = 3
        for (disc, prof, n_sem, n_prog, n_ced, pct) in tabelas[turma]:
            ws.cell(i, 1, disc)
            ws.cell(i, 2, "—" if prof in ("-", "NN") else
                    nomes_dos_profs(prof, siglas))
            ws.cell(i, 3, n_sem)
            ws.cell(i, 4, n_prog)
            ws.cell(i, 5, n_ced)
            cel = ws.cell(i, 6, pct if pct is not None else "—")
            if pct is not None:
                cel.number_format = "0.0%"
            for c in range(1, 7):
                x = ws.cell(i, c)
                x.border = borda
                x.alignment = Alignment(vertical="center", wrap_text=(c == 2),
                                        horizontal="center" if c != 2 else "left")
                if i % 2:
                    x.fill = zebra
            i += 1

        tp = sum(l[3] for l in tabelas[turma])
        tc = sum(l[4] for l in tabelas[turma])
        ws.cell(i, 1, "Total").font = Font(bold=True)
        ws.cell(i, 4, tp)
        ws.cell(i, 5, tc)
        cel = ws.cell(i, 6, (tc / tp) if tp else "—")
        if tp:
            cel.number_format = "0.0%"
        for c in range(1, 7):
            x = ws.cell(i, c)
            x.border, x.fill, x.font = borda, total_fill, Font(bold=True)
            x.alignment = Alignment(horizontal="center" if c != 2 else "left")

        for col, larg in zip("ABCDEF", (16, 46, 14, 18, 16, 14)):
            ws.column_dimensions[col].width = larg
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:F{i}"

    destino = os.path.join(OUT, "Relatorio_Tempos_Cedidos_1SEM.xlsx")
    wb.save(destino)
    return destino


if __name__ == "__main__":
    main()
