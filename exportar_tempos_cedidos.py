#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta o relatorio de tempos cedidos, por turma, a partir das propostas
ja geradas.

Para cada disciplina/professor da grade-base de cada turma, mostra quantas
aulas semanais ela tem, quantas aulas ela tem programadas no semestre
inteiro, quantas dessas aulas foram cedidas para a aplicacao de provas de
OUTRAS disciplinas (tempo emprestado) e que percentual da carga do
semestre isso representa.

Nao confia na memoria do gerador: rele as celulas de prova ja gravadas nas
planilhas e cruza, tempo a tempo, com a grade-base (GRADES) para descobrir
de quem era originalmente cada tempo usado por uma prova. Quando o tempo
nao e da propria disciplina examinada (nem, no caso de LP/LIT/RED, de uma
das tres disciplinas que compoem a prova combinada), conta como cedencia
do dono daquele tempo na grade.

Aulas programadas no semestre = nº de vezes que a disciplina/professor
aparece na grade semanal x nº de semanas letivas ativas da turma (da
semana 1 ate a ultima data de 2a chamada confirmada pela escola: 13/11
para as turmas 10/12, que viajam logo depois; 27/11 para as turmas 9/11),
descontando a semana vetada (12 a 16/10, conselho de classe: sem aula
normal) e os feriados que caem no mesmo dia da semana da aula.

Le  : Horario desenvolvido/Proposta_N_Calendario_Provas_2026_2SEM.xlsx
      siglas/siglas_profs_aux_etc.xlsx  (sigla -> nome do professor)
Grava: Horario desenvolvido/Relatorio_Tempos_Cedidos_Proposta_N.xlsx
       (uma aba por turma)
"""
import openpyxl, os, re, collections
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import gerar_calendario as G
from exportar_tabelas_turma import carregar_siglas, nomes_dos_profs

OUT = G.OUT
COLS = {"E": 1, "F": 2, "G": 3, "H": 4, "I": 5}
SIM_COD = re.compile(r"^(AG9|AG10|S\d-\d\d|EX\S+)")

# nome legivel -> codigo da grade (inverso de G.NOME; em caso de nomes
# repetidos - ex. "port"/"plit" -> "Port" -, fica o primeiro, que e o
# codigo realmente usado na grade)
REV_NOME = {}
for _cod, _nome in G.NOME.items():
    REV_NOME.setdefault(_nome, _cod)

# O calculo de aulas programadas no semestre vive no gerador (as regras de
# limite de cessao da Proposta 3 dependem dele), para nao haver duas
# versoes da mesma conta.
programadas_por_turma = G.programadas_no_semestre


def ler_exames(caminho, turma):
    """[(w, d, disc_nome, t_ini, n_tempos), ...] das celulas de prova.

    Simulados/AG ficam de fora: sao bloco fixo, nao passam pela logica de
    tempo emprestado.
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[turma]
    exames = []
    for w in range(1, 21):
        r = G.week_row(w)
        for col, d in COLS.items():
            v = ws[f"{col}{r}"].value
            if not v:
                continue
            txt = str(v).strip()
            if ("unterrichtsfrei" in txt or "Fach |" in txt or "Matéria |" in txt
                    or txt.startswith(("2CH", "CC", "Raumwunsch", "U-Stunden"))):
                continue
            partes = [p.strip() for p in txt.split("\n")]
            cabeca = partes[0]
            if SIM_COD.match(cabeca):
                continue
            disc = cabeca.split(" - ")[0].strip()
            ult = partes[-1] if len(partes) > 1 else ""
            nums = [int(x) for x in re.findall(r"(\d+)º", ult)]
            if not nums:
                continue
            t_ini = nums[0]
            n_t = (max(nums) - min(nums) + 1) if len(nums) >= 2 else 1
            exames.append((w, d, disc, t_ini, n_t))
    return exames


def cedencias_por_turma(caminho, turma):
    """{(codigo, prof): nº de aulas cedidas no semestre} da turma."""
    cont = collections.Counter()
    for (_w, d, disc, t_ini, n_t) in ler_exames(caminho, turma):
        if disc == "LP/LIT/RED":
            familia = G.COMBINA_PORT
        else:
            cod = REV_NOME.get(disc)
            if cod is None:
                print(f"  !! {turma}: disciplina não reconhecida '{disc}', "
                      "pulando na contagem de cedências")
                continue
            familia = {cod}
        for t in range(t_ini, t_ini + n_t):
            dono = G.GRADES[turma].get((d, t))
            if not dono:
                continue
            codigo, prof = dono
            if codigo in familia:
                continue          # tempo proprio da disciplina examinada
            cont[(codigo, prof)] += 1
    return cont


def exportar(proposta, siglas):
    origem = os.path.join(
        OUT, f"Proposta_{proposta}_Calendario_Provas_2026_2SEM.xlsx")
    destino = os.path.join(
        OUT, f"Relatorio_Tempos_Cedidos_Proposta_{proposta}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cab_fill = PatternFill("solid", fgColor="1F3864")
    cab_font = Font(bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    zebra = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    N_COL = 6

    for turma in G.GRADES:
        cedidas = cedencias_por_turma(origem, turma)
        semanais = collections.Counter(
            (codigo, prof) for (codigo, prof) in G.GRADES[turma].values())
        programadas = programadas_por_turma(turma)

        linhas = []
        for (codigo, prof), n_sem in semanais.items():
            n_prog = programadas[(codigo, prof)]
            n_ced = cedidas.get((codigo, prof), 0)
            pct = (n_ced / n_prog) if n_prog else None
            linhas.append((G.NOME.get(codigo, codigo), prof, n_sem, n_prog,
                           n_ced, pct))
        linhas.sort(key=lambda x: (-x[4], x[0], x[1]))

        ws = wb.create_sheet(turma)
        ws["A1"] = (f"Tempos Cedidos para Provas — {turma} — "
                    f"2º semestre 2026 (Proposta {proposta})")
        ws["A1"].font = Font(bold=True, size=13)
        ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(N_COL)}1")
        ws["A1"].alignment = Alignment(horizontal="center")

        cabecalhos = ["Disciplina", "Professor(es)", "Nº de aulas semanais",
                      "Nº de aulas programadas no semestre",
                      "Nº de aulas cedidas para provas de outras disciplinas",
                      "% de aulas cedidas no semestre"]
        for c, t in enumerate(cabecalhos, 1):
            cel = ws.cell(2, c, t)
            cel.fill, cel.font = cab_fill, cab_font
            cel.alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=True)
            cel.border = borda

        i = 3
        for (disc, prof, n_sem, n_prog, n_ced, pct) in linhas:
            prof_txt = "—" if prof == "-" else nomes_dos_profs(prof, siglas)
            ws.cell(i, 1, disc)
            ws.cell(i, 2, prof_txt)
            ws.cell(i, 3, n_sem)
            ws.cell(i, 4, n_prog)
            ws.cell(i, 5, n_ced)
            cel_pct = ws.cell(i, 6, pct if pct is not None else "—")
            if pct is not None:
                cel_pct.number_format = "0.0%"
            for c in range(1, N_COL + 1):
                cel = ws.cell(i, c)
                cel.border = borda
                cel.alignment = Alignment(
                    vertical="center", wrap_text=(c == 2),
                    horizontal="center" if c != 2 else "left")
                if i % 2:
                    cel.fill = zebra
            i += 1

        tot_prog = sum(x[3] for x in linhas)
        tot_ced = sum(x[4] for x in linhas)
        ws.cell(i, 1, "Total")
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 4, tot_prog)
        ws.cell(i, 5, tot_ced)
        cel_tot_pct = ws.cell(i, 6, (tot_ced / tot_prog) if tot_prog else "—")
        if tot_prog:
            cel_tot_pct.number_format = "0.0%"
        for c in range(1, N_COL + 1):
            cel = ws.cell(i, c)
            cel.border = borda
            cel.fill = total_fill
            cel.font = Font(bold=True)
            cel.alignment = Alignment(horizontal="center" if c != 2 else "left")

        for col, larg in zip("ABCDEF", (16, 46, 14, 18, 16, 14)):
            ws.column_dimensions[col].width = larg
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(N_COL)}{i}"

    wb.save(destino)
    return destino


def main():
    siglas = carregar_siglas()
    for p in (3,):   # so a Proposta 3 e desenvolvida (pedido do usuario)
        print("Gerado:", exportar(p, siglas))


if __name__ == "__main__":
    main()
