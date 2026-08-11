#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Exporta uma tabela única com todas as provas, organizada por professor,
a partir da Proposta já gravada.

Diferença para Tabela_Provas_por_Turma: aquela é 1 aba por turma; esta é
1 tabela só, para o professor achar de uma vez todas as provas que ele
precisa acompanhar -- inclusive quando dá aula em mais de uma turma.
Quando a mesma prova (mesma disciplina/data/tempos) é aplicada em mais de
uma turma simultaneamente (professor comum entre turmas irmãs), vira uma
linha só com as turmas juntas na coluna "Turma(s)" em vez de duplicar.

Não inclui simulados/AG: não têm um professor especifico responsável (são
aplicados por vários fiscais), então não fazem sentido numa tabela "por
professor".

Colunas: Professor | Disciplina | Data | Dia da semana | Tempos | Turma(s)

Lê  : Horario desenvolvido/Proposta_N_Calendario_Provas_2026_2SEM.xlsx
      siglas/siglas_profs_aux_etc.xlsx  (sigla -> nome do professor)
Grava: Horario desenvolvido/Provas_por_Professor_Proposta_N.xlsx
"""
import openpyxl, os, re, datetime, collections
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import gerar_calendario as G
from exportar_tabelas_turma import carregar_siglas

BASE = G.BASE
OUT = G.OUT
SEMANA1 = datetime.date(2026, 8, 3)
COLS = {"E": 1, "F": 2, "G": 3, "H": 4, "I": 5}
DIA_NOME = {1: "Segunda", 2: "Terça", 3: "Quarta", 4: "Quinta", 5: "Sexta"}
SIM_COD = re.compile(r"^(AG9|AG10|S\d-\d\d|EX\S+)")


def data_de(w, d):
    return SEMANA1 + datetime.timedelta(days=7 * (w - 1) + d - 1)


def nome_do(sigla, siglas):
    nome = siglas.get(sigla)
    if not nome:
        for k, v in siglas.items():
            if k.lower() == sigla.lower():
                nome = v
                break
    return f"{sigla} ({nome})" if nome else sigla


def ler_provas_por_professor(caminho, turma):
    """[(sigla, disc, data, tempos_txt, turma), ...] -- uma linha por
    professor CITADO na prova (uma disciplina com professor comum entre
    turmas irmãs / grupo paralelo gera uma linha por sigla)."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[turma]
    linhas = []
    for w in range(1, 21):
        r = G.week_row(w)
        for col, d in COLS.items():
            v = ws[f"{col}{r}"].value
            if not v:
                continue
            txt = str(v).strip()
            if ("unterrichtsfrei" in txt or "Fach |" in txt or "Matéria |" in txt
                    or txt.startswith(("2CH", "CC", "Raumwunsch", "U-Stunden"))):
                continue
            partes = [p.strip() for p in txt.split("\n")]
            cabeca = partes[0]
            tempos_txt = partes[-1] if len(partes) > 1 else ""
            if SIM_COD.match(cabeca) or " - " not in cabeca:
                continue
            disc, campo = cabeca.split(" - ", 1)
            disc = disc.strip()
            dt = data_de(w, d)
            for sigla in (p.strip() for p in re.split(r"[/\-]", campo)):
                if sigla:
                    linhas.append((sigla, disc, dt, tempos_txt, turma))
    return linhas


def montar_linhas(caminho, siglas):
    brutas = []
    for turma in G.GRADES:
        brutas += ler_provas_por_professor(caminho, turma)

    # junta turmas quando a mesma prova (professor+disciplina+data+tempos)
    # aparece em mais de uma turma ao mesmo tempo (professor comum)
    agrupado = collections.OrderedDict()
    for (sigla, disc, dt, tempos, turma) in brutas:
        chave = (sigla, disc, dt, tempos)
        agrupado.setdefault(chave, set()).add(turma)

    linhas = []
    for (sigla, disc, dt, tempos), turmas in agrupado.items():
        nome_completo = nome_do(sigla, siglas)
        linhas.append((nome_completo, disc, dt, tempos, ", ".join(sorted(turmas))))

    linhas.sort(key=lambda x: (x[0], x[2]))
    return linhas


def exportar(proposta, siglas):
    origem = os.path.join(
        OUT, f"Proposta_{proposta}_Calendario_Provas_2026_2SEM.xlsx")
    destino = os.path.join(
        OUT, f"Provas_por_Professor_Proposta_{proposta}.xlsx")

    linhas = montar_linhas(origem, siglas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provas por professor"

    cab_fill = PatternFill("solid", fgColor="1F3864")
    cab_font = Font(bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    zebra = PatternFill("solid", fgColor="F2F2F2")
    N_COL = 6

    ws["A1"] = f"Provas por Professor — 2º semestre 2026 (Proposta {proposta})"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(N_COL)}1")
    ws["A1"].alignment = Alignment(horizontal="center")

    cabecalhos = ["Professor", "Disciplina", "Data", "Dia da semana",
                  "Tempos", "Turma(s)"]
    for c, t in enumerate(cabecalhos, 1):
        cel = ws.cell(2, c, t)
        cel.fill, cel.font = cab_fill, cab_font
        cel.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        cel.border = borda

    i = 3
    for (professor, disc, dt, tempos, turmas) in linhas:
        ws.cell(i, 1, professor)
        ws.cell(i, 2, disc)
        cel_data = ws.cell(i, 3, dt)
        cel_data.number_format = "DD/MM/YYYY"
        ws.cell(i, 4, DIA_NOME[dt.weekday() + 1])
        ws.cell(i, 5, tempos)
        ws.cell(i, 6, turmas)
        for c in range(1, N_COL + 1):
            cel = ws.cell(i, c)
            cel.border = borda
            cel.alignment = Alignment(
                vertical="center", wrap_text=(c == 1),
                horizontal="left" if c in (1, 2) else "center")
            if i % 2:
                cel.fill = zebra
        i += 1

    for col, larg in zip("ABCDEF", (34, 16, 12, 13, 30, 14)):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(N_COL)}{i - 1}"

    wb.save(destino)
    return destino


def main():
    siglas = carregar_siglas()
    for p in (3,):   # so a Proposta 3 e desenvolvida (pedido do usuario)
        print("Gerado:", exportar(p, siglas))


if __name__ == "__main__":
    main()
