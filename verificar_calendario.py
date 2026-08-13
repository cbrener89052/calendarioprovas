#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Confere as planilhas geradas contra o checklist da skill calendario-provas.

Le os arquivos .xlsx ja gravados (nao confia na memoria do gerador).
"""
import openpyxl, os, re, collections, datetime
import gerar_calendario as G

OUT = G.OUT
COLS = {"E": 1, "F": 2, "G": 3, "H": 4, "I": 5}

LIMITE = {"10_12": datetime.date(2026, 11, 12),
          "9_11": datetime.date(2026, 11, 26)}
INICIO_P1 = datetime.date(2026, 8, 17)
FERIADOS = {datetime.date(2026, 9, 7), datetime.date(2026, 11, 20)}
SEMANA_VETADA = (datetime.date(2026, 10, 12), datetime.date(2026, 10, 16))

SIM_COD = re.compile(r"^(AG9|AG10|S\d-\d\d|EX\S+|DSD\d+)")
# subconjunto reconhecido pela regra oficial de simulados (G.SIMULADOS) --
# DSD fica de fora de proposito: e uma avaliacao externa pontual (pedido
# do usuario, 08/2026: "nao colocar na regra"), nao um simulado recorrente
# do calendario gerado por gerar_calendario.py.
SIM_COD_OFICIAL = re.compile(r"^(AG9|AG10|S\d-\d\d|EX\S+)")

# excecoes a regra 11 (provas de professor comum entre turmas irmas
# precisam coincidir em dia/tempo): pares (turma_a, turma_b, nome da
# disciplina em G.NOME) que ficam de fora dessa exigencia.
# Regra fixa desta escola (pedido do usuario, 08/2026: "os professores
# de filosofia e de sociologia aplicam provas nos seus tempos de aula e
# nao simultaneamente em turmas irmas"): Filosofia e Sociologia, mesmo
# com professor comum entre turmas irmas, NUNCA precisam coincidir --
# cada turma usa o tempo proprio do professor nela. Nao vale para
# nenhuma outra disciplina com professor comum (essas continuam exigindo
# coincidencia, ver skill). Ao configurar um novo semestre, adicionar
# aqui automaticamente todo par turma-irma cuja disciplina comum seja
# Filosofia ou Sociologia.
COORDENACAO_EXCECAO = {
    ("10C1", "10C2", "Soc"),
    ("12C1", "12C2", "Soc"),
    ("10C1", "10C2", "Fil"),
    ("11C1", "11C2", "Fil"),
    ("12C1", "12C2", "Fil"),
}


SEMANA1 = datetime.date(2026, 8, 3)


def data_da(ws, w, d):
    """Semana 1 comeca em 03/08/2026 (as celulas de data sao formulas)."""
    return SEMANA1 + datetime.timedelta(days=7 * (w - 1) + d - 1)


def main():
    problemas = []
    avisos = []
    G.carregar_ocupadas()  # preenche G.LIMITE_LPLITRED_CONSELHO por grupo de turma
    comuns_por_par = {(a, b): G.classificar_par(a, b) for a, b in G.PARES_IRMAS}
    irma = {a: b for a, b in G.PARES_IRMAS}
    irma.update({b: a for a, b in G.PARES_IRMAS})

    for prop in (3,):   # so a Proposta 3 e desenvolvida (pedido do usuario)
        caminho = os.path.join(OUT, f"Proposta_{prop}_Calendario_Provas_2026_2SEM.xlsx")
        wb = openpyxl.load_workbook(caminho, data_only=True)
        provas_por_turma = {}
        for turma in G.GRADES:
            ws = wb[turma]
            provas = []          # (data, w, d, disc, texto, tempo_ini, n_tempos, destacada)
            for w in range(1, 21):
                r = G.week_row(w)
                for col, d in COLS.items():
                    v = ws[f"{col}{r}"].value
                    if not v:
                        continue
                    txt = str(v).strip()
                    if "unterrichtsfrei" in txt or "Fach |" in txt or \
                       "Matéria |" in txt or txt.startswith(("2CH", "CC")):
                        continue
                    disc = txt.split("\n")[0].split(" - ")[0].strip()
                    ult = txt.split("\n")[-1]
                    nums = [int(x) for x in re.findall(r"(\d+)º", ult)]
                    t_ini = nums[0] if nums else None
                    n_t = (max(nums) - min(nums) + 1) if len(nums) >= 2 else \
                        (1 if nums else None)
                    cell = ws[f"{col}{r}"]
                    cor = getattr(cell.fill.start_color, "rgb", None)
                    destacada = cor not in (None, "00000000")
                    provas.append((data_da(ws, w, d), w, d, disc, txt,
                                  t_ini, n_t, destacada))
            provas_por_turma[turma] = provas

        for turma in G.GRADES:
            provas = provas_por_turma[turma]
            pre = f"P{prop}/{turma}"
            grupo = G.grupo_turma(turma)

            # 0. REGRA PRIORIDADE 1: o professor tem que estar presente na
            #    aplicacao da propria prova (aula propria da familia da
            #    disciplina, na propria turma OU na turma irma), em pelo
            #    menos 1 tempo do bloco de aplicacao -- nao precisa ter os
            #    2 tempos. Quando a prova cita mais de um professor (grupo
            #    paralelo / professor comum), basta 1 deles bater.
            for (_dt, w, d, disc, txt, t_ini, n_t, _ds) in provas:
                if SIM_COD.match(disc) or t_ini is None or n_t is None:
                    continue
                cabeca = txt.split("\n")[0]
                if " - " not in cabeca:
                    continue
                prof_txt = cabeca.split(" - ", 1)[1].strip()
                chave = [k for k, v in G.NOME.items() if v == disc]
                if not chave:
                    continue
                if not G.professor_presente_no_bloco(
                        turma, chave[0], prof_txt, d, t_ini, n_t):
                    problemas.append(
                        f"{pre}: semana {w} {disc}/{prof_txt} -- nenhum "
                        f"professor citado tem aula própria (turma ou "
                        f"turma irmã) nesse dia/bloco (tempos "
                        f"{t_ini}-{t_ini + n_t - 1})")

            # 1. max 3 avaliacoes por semana (simulado de 2 dias conta 1x)
            por_sem = collections.defaultdict(set)
            for (dt, w, d, disc, _tx, _ti, _nt, _ds) in provas:
                cod = SIM_COD.match(disc)
                por_sem[w].add(cod.group(1) if cod else disc)
            for w, s in por_sem.items():
                if len(s) > 3:
                    problemas.append(f"{pre}: semana {w} com {len(s)} avaliações {sorted(s)}")

            # 2. grupo 1 nao coincide. Excecao confirmada pela escola: duas
            #    do grupo 1 na mesma semana sao aceitas quando uma delas e
            #    Ingles (nunca tres; nunca um par sem Ingles).
            for w, s in por_sem.items():
                g1 = [x for x in s if x.lower() in
                      {"mat", "daf", "port", "lp/lit/red", "ing"}]
                if len(g1) > 2 or (len(g1) == 2 and
                                   not any(x.lower() == "ing" for x in g1)):
                    problemas.append(f"{pre}: semana {w} com grupo 1 repetido {g1}")

            # 3. um exame por dia
            dias = collections.Counter((w, d) for (_, w, d, _tx, _x, _ti, _nt, _ds) in provas)
            for k, c in dias.items():
                if c > 1:
                    problemas.append(f"{pre}: {c} provas no mesmo dia {k}")

            # 4. datas: dentro do periodo, sem feriado, sem a semana vetada
            for (dt, w, d, disc, _tx, _ti, _nt, _ds) in provas:
                if dt < INICIO_P1:
                    problemas.append(f"{pre}: {disc} em {dt} antes de 17/08")
                if dt > LIMITE[grupo]:
                    problemas.append(f"{pre}: {disc} em {dt} depois do limite "
                                     f"{LIMITE[grupo]}")
                if dt in FERIADOS:
                    problemas.append(f"{pre}: {disc} em feriado {dt}")
                if SEMANA_VETADA[0] <= dt <= SEMANA_VETADA[1]:
                    problemas.append(f"{pre}: {disc} na semana vetada {dt}")

            # 5. numero de provas por disciplina
            cont = collections.Counter(
                disc for (_, _, _, disc, _tx, _ti, _nt, _ds) in provas if not SIM_COD.match(disc))
            for disc, c in cont.items():
                _chave = [k for k, v in G.NOME.items() if v == disc]
                _codigo = _chave[0] if _chave else None
                esperado = 1 if disc in ("Fil", "Soc") or \
                    (turma.startswith("9") and disc in ("Fis", "Qui")) or \
                    (_codigo and _codigo in G.UMA_PROVA_POR_TURMA.get(turma, ())) else 2
                if disc in ("Port", "Redação", "Gram") and not turma.startswith("9"):
                    problemas.append(f"{pre}: {disc} separado (deveria ser LP/LIT/RED)")
                if c != esperado:
                    problemas.append(f"{pre}: {disc} tem {c} provas (esperado {esperado})")

            # 5a. distancia minima entre as 2 provas do semestre da mesma
            #     disciplina (disciplinas de prova unica nao entram, ja
            #     tem so 1 ocorrencia)
            por_disc = collections.defaultdict(list)
            for (_, w, _, disc, _tx, _ti, _nt, _ds) in provas:
                if not SIM_COD.match(disc):
                    por_disc[disc].append(w)
            for disc, semanas in por_disc.items():
                if len(semanas) == 2:
                    dist = abs(semanas[0] - semanas[1])
                    if dist < G.DISTANCIA_MIN_MESMA_DISC:
                        problemas.append(
                            f"{pre}: {disc} com só {dist} semana(s) entre as "
                            f"2 provas (mínimo {G.DISTANCIA_MIN_MESMA_DISC})")

            # 5a-bis. LP/LIT/RED precisa cair ate 10 dias corridos antes do
            # conselho de classe relevante pro grupo de turma (ver skill)
            limite_lplitred = G.LIMITE_LPLITRED_CONSELHO.get(
                G.grupo_turma(turma), G.LIMITE_LPLITRED_CONSELHO_BASE)
            for w in por_disc.get("LP/LIT/RED", ()):
                if w > limite_lplitred:
                    problemas.append(
                        f"{pre}: LP/LIT/RED na semana {w}, depois do limite "
                        f"de {limite_lplitred} semanas antes do conselho de "
                        f"classe relevante para essa turma")

            # 5b. LP/LIT/RED usa 3 tempos nas turmas 10/11/12
            for (_, _, _, disc, txt, _ti, _nt, _ds) in provas:
                if disc == "LP/LIT/RED":
                    ult = txt.split("\n")[-1]
                    if "ao" not in ult:
                        problemas.append(f"{pre}: LP/LIT/RED sem 3 tempos ({ult})")

            # 5c. simulados do 2o ao 7o tempo
            for (_, _, _, disc, txt, _ti, _nt, _ds) in provas:
                if SIM_COD.match(disc) and "2º ao 7º" not in txt:
                    problemas.append(f"{pre}: simulado {disc} fora do 2o-7o tempo")

            # 6. disciplinas de 1 tempo usam 1 tempo
            for (_, _, _, disc, txt, _ti, _nt, _ds) in provas:
                um_tempo = disc in ("Fil", "Soc") or \
                    (turma.startswith("9") and disc in ("Fis", "Qui"))
                if um_tempo and " e " in txt.split("\n")[-1]:
                    problemas.append(f"{pre}: {disc} com 2 tempos ({txt.split(chr(10))[-1]})")

            # 7. Ed. Fisica e afins nao tem prova
            for (_, _, _, disc, _tx, _ti, _nt, _ds) in provas:
                if disc.lower() in ("esp", "art", "tec", "finan", "socem", "proj"):
                    problemas.append(f"{pre}: prova indevida de {disc}")

            # 7b. provas nos tempos 7-11 so quando forem inevitaveis:
            #     a disciplina precisa nao ter NENHUM slot pela manha
            for (_, _, _, disc, txt, _ti, _nt, _ds) in provas:
                if SIM_COD.match(disc):
                    continue
                ult = txt.split("\n")[-1]
                nums = [int(x) for x in re.findall(r"(\d+)º", ult)]
                if not nums or max(nums) < G.PRIMEIRO_TEMPO_TARDE:
                    continue
                chave = [k for k, v in G.NOME.items() if v == disc]
                if not chave:
                    continue
                d0 = chave[0]
                n = len(nums) if len(nums) > 1 else 1
                if d0 == "LPLITRED":
                    n = 3
                elif len(nums) == 2:
                    n = 2
                # opcao de manha valida = nao e tarde E nao cruza o intervalo
                # do recreio (essa restricao tem prioridade sobre a da tarde)
                cedo = {(dd, tt) for (dd, tt, _x) in
                        G.slots_da_disciplina(turma, d0, n)
                        if not G._tarde(tt, n) and not G.cruza_intervalo(tt, n)}
                # prova de professor comum e aplicada nas duas turmas irmas ao
                # mesmo tempo: so vale como "opcao de manha" o slot que existe
                # nas DUAS grades
                par = comuns_por_par.get(tuple(sorted((turma, irma.get(turma, turma)))), {})
                if d0 in par:
                    outra = irma[turma]
                    cedo &= {(dd, tt) for (dd, tt, _x) in
                             G.slots_da_disciplina(outra, d0, n)
                             if not G._tarde(tt, n) and not G.cruza_intervalo(tt, n)}
                if cedo:
                    # na Proposta 3 os tetos de cessao tiram opcoes de manha
                    # que este calculo (que so olha a grade) ainda enxerga —
                    # por isso vira aviso, nao falha
                    destino = avisos if prop == 3 else problemas
                    destino.append(f"{pre}: {disc} nos tempos 7-11 ({ult}) "
                                   f"tendo {len(cedo)} opcao(oes) de manha")

            # 7c. disciplina de 1 tempo semanal nao pode ter cedido tempo
            veto = G.nao_doadoras(turma)
            for (_, _, d_prova, disc, txt, _ti, _nt, _ds) in provas:
                if SIM_COD.match(disc):
                    continue
                chave = [k for k, v in G.NOME.items() if v == disc]
                if not chave:
                    continue
                d0 = chave[0]
                ult = txt.split("\n")[-1]
                nums = [int(x) for x in re.findall(r"(\d+)º", ult)]
                if len(nums) < 2:
                    continue
                faixa = range(min(nums), max(nums) + 1)
                # numa prova combinada (LP/LIT/RED) o tempo de Gramática ou
                # Redação é tempo PRÓPRIO da prova, não uma cessão
                familia = G.familia_de(d0)
                for t in faixa:
                    dono = G.GRADES[turma].get((d_prova, t))
                    if dono and dono[0] not in familia and dono[0] in veto:
                        problemas.append(
                            f"{pre}: {disc} usa o {t}º tempo de {dono[0]}, "
                            f"que so tem 1 aula na semana")

            # 8. simulados nas datas oficiais
            oficiais = {(w, d) for (w, d, c, l) in G.SIMULADOS[turma]}
            achados = {(w, d) for (_, w, d, disc, _tx, _ti, _nt, _ds) in provas if SIM_COD_OFICIAL.match(disc)}
            if oficiais != achados:
                problemas.append(f"{pre}: simulados {sorted(achados)} != "
                                 f"oficiais {sorted(oficiais)}")

            # 9. nenhuma prova cruza o intervalo do recreio, salvo as
            #    sinalizadas com destaque de cor na celula
            for (_, w, d, disc, txt, t_ini, n_t, destacada) in provas:
                if SIM_COD.match(disc) or t_ini is None or n_t is None:
                    continue
                if G.cruza_intervalo(t_ini, n_t) and not destacada:
                    problemas.append(
                        f"{pre}: {disc} cruza o intervalo do recreio "
                        f"({txt.split(chr(10))[-1]}) sem destaque na célula")

            # 10. datas exigidas pela coordenacao (FORCAR_DATA)
            for (turma2, disc2, _per2), (w_f, d_f) in G.FORCAR_DATA.items():
                if turma2 != turma:
                    continue
                nome2 = G.NOME.get(disc2, disc2)
                bate = [1 for (_, w, d, disc, *_r) in provas
                        if disc == nome2 and (w, d) == (w_f, d_f)]
                if not bate:
                    problemas.append(
                        f"{pre}: {nome2} não está na data exigida "
                        f"(semana {w_f}, dia {d_f})")

        # 10b. Proposta 3: limites de cessao de aula (regras 1 a 5 pedidas
        #      pela coordenacao). Nao valem para as Propostas 1 e 2.
        if prop == 3:
            for turma in G.GRADES:
                pre = f"P{prop}/{turma}"
                cedidas = collections.defaultdict(list)   # (disc,prof) -> [(w,d,t)]
                exames_de = collections.defaultdict(set)  # disc -> {semana}
                for (_dt, w, d, disc, _tx, t_ini, n_t, _ds) in provas_por_turma[turma]:
                    if SIM_COD.match(disc) or t_ini is None or n_t is None:
                        continue
                    chave = [k for k, v in G.NOME.items() if v == disc]
                    fam = G.familia_de(chave[0]) if chave else set()
                    for c in fam:
                        exames_de[c].add(w)
                    for t in range(t_ini, t_ini + n_t):
                        dono = G.GRADES[turma].get((d, t))
                        if not dono or dono[0] in fam:
                            continue
                        cedidas[dono].append((w, d, t))

                semanais = G.aulas_semanais(turma)
                prog = G.programadas_no_semestre(turma)
                for (chave, slots) in cedidas.items():
                    disc_d, prof_d = chave
                    nome_d = G.NOME.get(disc_d, disc_d)
                    n_sem = semanais.get(chave, 0)
                    n_ced = len(slots)

                    # regra 2: 1 aula por semana nao cede -- e a unica das
                    # cinco que o gerador NUNCA relaxa (limite_cessao
                    # devolve 0 sempre, folga nao entra nessa conta), entao
                    # continua PROBLEMA se acontecer.
                    if n_sem <= 1:
                        problemas.append(
                            f"{pre}: {nome_d}/{prof_d} tem 1 aula semanal e "
                            f"cedeu {n_ced}")
                    # regra 1: 2 ou 3 aulas semanais -> meta de 2 (3 para
                    # Historia, Geografia e GL, por decisão da escola). Junto
                    # com a regra 5 abaixo, pode ser relaxada POR TURMA (o
                    # teto de cessões elevado, ver escada em main()) quando a
                    # busca estrita nao fecha -- o gerador avisa, entao entra
                    # como AVISO, nao como falha do checklist (mesmo
                    # tratamento ja dado a regra 4 logo abaixo).
                    else:
                        meta = G.meta_cessao(disc_d, n_sem)
                        if meta is not None and n_ced > meta:
                            avisos.append(
                                f"{pre}: {nome_d}/{prof_d} tem {n_sem} aulas "
                                f"semanais e cedeu {n_ced} (meta: no máximo "
                                f"{meta})")
                    # regra 5: teto percentual das aulas do semestre
                    n_prog = prog.get(chave, 0)
                    if n_prog and n_ced / n_prog > G.TETO_PCT_CESSAO:
                        avisos.append(
                            f"{pre}: {nome_d}/{prof_d} cedeu {n_ced} de "
                            f"{n_prog} aulas ({n_ced / n_prog:.1%}) — acima "
                            f"do teto de {G.TETO_PCT_CESSAO:.0%}")
                    # regra 4: nao ceder as vesperas da propria prova. Foi
                    # relaxada por inviabilidade (o gerador avisa), entao
                    # entra como AVISO, nao como falha do checklist.
                    for (w, _d, _t) in slots:
                        for we in exames_de.get(disc_d, ()):
                            if we in (w, w + 1):
                                avisos.append(
                                    f"{pre}: {nome_d}/{prof_d} cedeu aula na "
                                    f"semana {w} tendo prova própria na "
                                    f"semana {we}")
                    # regra 3: duas semanas seguidas sem contato com a turma
                    c = G.Cessoes(turma)
                    antes = c._pares_sem_contato(chave)
                    for (w, _d, _t) in slots:
                        c.cedidas_sem[chave][w] += 1
                    depois = c._pares_sem_contato(chave)
                    for w in sorted(depois - antes):
                        problemas.append(
                            f"{pre}: {nome_d}/{prof_d} fica sem contato com a "
                            f"turma nas semanas {w} e {w + 1} por causa das "
                            f"cessões")

        # 11. provas de professor comum entre turmas irmas: mesmo dia/tempo
        for (a, b), comuns in comuns_por_par.items():
            pa = {}
            for (_, w, d, disc, txt, t_ini, n_t, _ds) in provas_por_turma[a]:
                pa.setdefault(disc, []).append((w, d, t_ini, n_t))
            pb = {}
            for (_, w, d, disc, txt, t_ini, n_t, _ds) in provas_por_turma[b]:
                pb.setdefault(disc, []).append((w, d, t_ini, n_t))
            for disc, (_tipo, _profs) in comuns.items():
                nome = G.NOME.get(disc, disc)
                if (a, b, nome) in COORDENACAO_EXCECAO:
                    continue
                if sorted(pa.get(nome, [])) != sorted(pb.get(nome, [])):
                    problemas.append(
                        f"P{prop}/{a}-{b}: {nome} tem professor comum mas as "
                        f"provas não coincidem: {sorted(pa.get(nome, []))} vs "
                        f"{sorted(pb.get(nome, []))}")

    if problemas:
        print(f"{len(problemas)} PROBLEMA(S):")
        for p in problemas:
            print("  -", p)
    else:
        print("OK: as propostas passaram em todos os itens do checklist.")
    if avisos:
        print(f"\n{len(avisos)} AVISO(S) — regras relaxadas por inviabilidade, "
              "não são falhas do checklist:")
        for a in avisos:
            print("  ~", a)


if __name__ == "__main__":
    main()
