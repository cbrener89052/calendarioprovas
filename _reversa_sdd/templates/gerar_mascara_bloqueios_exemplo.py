#!/usr/bin/env python3
"""Gera Mascara_Bloqueios_Calendario_2026_2SEM.xlsx a partir do legado.

Fontes: gerar_calendario.py (BLOQUEIOS, SEMANA_BLOQUEADA, SIMULADOS,
FORCAR_DATA) e verificar_calendario.py (FERIADOS).
"""
from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

SEMANA1 = datetime.date(2026, 8, 3)

FERIADOS = [
    (datetime.date(2026, 9, 7), "Independência do Brasil", "feriado"),
    (datetime.date(2026, 11, 2), "Finados", "feriado"),
    (datetime.date(2026, 11, 20), "Consciência Negra", "feriado"),
]

SEMANAS_VETADAS = [
    (11, "Conselho de classe / semana 12-16 out (N. Sra. Aparecida)", ""),
]

# BLOQUEIOS legado — espelhados em feriados por data; aba dias_bloqueados
# fica reservada a bloqueios operacionais não-feriado.
DIAS_BLOQUEADOS: list[tuple] = []

SIMULADOS = {
    "9C1": [(9, 5, "AG9", "2º ao 7º tempos")],
    "9C2": [(9, 5, "AG9", "2º ao 7º tempos")],
    "10C1": [(8, 5, "AG10", "2º ao 7º tempos")],
    "10C2": [(8, 5, "AG10", "2º ao 7º tempos")],
    "11C1": [
        (4, 2, "S3-11", "2º ao 7º tempos"),
        (4, 3, "S3-11", "2º ao 7º tempos"),
        (13, 1, "S4-11", "2º ao 7º tempos"),
        (13, 2, "S4-11", "2º ao 7º tempos"),
    ],
    "11C2": [
        (4, 2, "S3-11", "2º ao 7º tempos"),
        (4, 3, "S3-11", "2º ao 7º tempos"),
        (13, 1, "S4-11", "2º ao 7º tempos"),
        (13, 2, "S4-11", "2º ao 7º tempos"),
    ],
    "12C1": [
        (8, 3, "S4-12", "2º ao 7º tempos"),
        (8, 4, "S4-12", "2º ao 7º tempos"),
    ],
    "12C2": [
        (8, 3, "S4-12", "2º ao 7º tempos"),
        (8, 4, "S4-12", "2º ao 7º tempos"),
    ],
}

FORCAR_DATA = [
    ("10C2", "ing", 1, 5, 2, "1ª prova Inglês — pedido coordenação 01/09/2026"),
]

OUT = Path(__file__).with_name("Mascara_Bloqueios_Calendario_2026_2SEM.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)


def semana_dia_para_data(semana: int, dia: int) -> datetime.date:
    return SEMANA1 + datetime.timedelta(days=7 * (semana - 1) + (dia - 1))


def escrever_cabecalho(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def gerar() -> Path:
    wb = openpyxl.Workbook()

    # --- feriados ---
    ws = wb.active
    ws.title = "feriados"
    escrever_cabecalho(ws, ["data", "descricao", "tipo", "escopo"])
    for data, desc, tipo in FERIADOS:
        ws.append([data.isoformat(), desc, tipo, ""])

    # --- semanas_vetadas ---
    ws = wb.create_sheet("semanas_vetadas")
    escrever_cabecalho(ws, ["semana", "motivo", "escopo"])
    for semana, motivo, escopo in SEMANAS_VETADAS:
        ws.append([semana, motivo, escopo])

    # --- dias_bloqueados ---
    ws = wb.create_sheet("dias_bloqueados")
    escrever_cabecalho(ws, ["semana", "dia", "motivo", "turmas", "bloqueia_prova"])
    for semana, dia, motivo, turmas in DIAS_BLOQUEADOS:
        ws.append([semana, dia, motivo, turmas, "S"])

    # --- simulados ---
    ws = wb.create_sheet("simulados")
    escrever_cabecalho(ws, ["turma", "semana", "dia", "codigo", "tempos", "observacao"])
    for turma in sorted(SIMULADOS):
        for semana, dia, codigo, tempos in SIMULADOS[turma]:
            data = semana_dia_para_data(semana, dia)
            obs = f"Legado gerar_calendario.py — {data.strftime('%d/%m/%Y')}"
            ws.append([turma, semana, dia, codigo, tempos, obs])

    # --- datas_forcadas ---
    ws = wb.create_sheet("datas_forcadas")
    escrever_cabecalho(ws, ["turma", "disciplina", "periodo", "semana", "dia", "motivo"])
    for turma, disc, periodo, semana, dia, motivo in FORCAR_DATA:
        ws.append([turma, disc, periodo, semana, dia, motivo])

    # --- metadados ---
    ws = wb.create_sheet("_meta")
    ws.append(["campo", "valor"])
    ws.append(["semestre", "2026 — 2º semestre"])
    ws.append(["SEMANA1", SEMANA1.isoformat()])
    ws.append(["fonte", "gerar_calendario.py + verificar_calendario.py"])
    ws.append(["spec", "_reversa_sdd/templates/mascara-bloqueios-calendario-spec.md"])
    ws.sheet_state = "hidden"

    for sheet in wb.worksheets:
        if sheet.title == "_meta":
            continue
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = max(
                12, max(len(str(c.value or "")) for c in col) + 2
            )

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = gerar()
    print(f"Gerado: {path}")
