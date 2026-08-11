#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera o relatório de trocas de tempo entre professores a partir da
planilha FINAL da Proposta 3 (não confia na memória do gerador nem exige
rodar o solver de novo -- por isso é seguro de rodar a qualquer momento,
mesmo depois de edições manuais na Proposta 3, sem risco de sobrescrever
o calendário já fechado).

Grava dois formatos, sempre a partir dos mesmos dados:
- Horario desenvolvido/Relatorio_trocas_de_tempo.md
  (reusa gerar_calendario.relatorio(), alimentada com os itens
  (w, d, t, n, disc, prof, doador) reconstruídos da planilha gravada, em
  vez das alocações do backtracking)
- Horario desenvolvido/Relatorio_trocas_de_tempo.xlsx
  (mesmo conteúdo, em 3 abas, com nome completo dos professores -- para
  compartilhar diretamente com a coordenação/professores)
"""
import os, re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import gerar_calendario as G
from exportar_tabelas_turma import carregar_siglas, nomes_dos_profs

OUT = G.OUT
COLS = {"E": 1, "F": 2, "G": 3, "H": 4, "I": 5}
SIM_COD = re.compile(r"^(AG9|AG10|S\d-\d\d|EX\S+)")


def ler_itens(caminho, turma):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[turma]
    itens = []
    for w in range(1, 21):
        r = G.week_row(w)
        for col, d in COLS.items():
            v = ws[f"{col}{r}"].value
            if not v:
                continue
            txt = str(v).strip()
            if ("unterrichtsfrei" in txt or "Fach |" in txt or "Matéria |" in txt
                    or txt.startswith(("2CH", "CC", "Raumwunsch", "U-Stunden"))):
                continue
            cabeca = txt.split("\n")[0]
            if SIM_COD.match(cabeca) or " - " not in cabeca:
                continue
            disc_nome, prof_txt = cabeca.split(" - ", 1)
            disc_nome, prof_txt = disc_nome.strip(), prof_txt.strip()
            ult = txt.split("\n")[-1]
            nums = [int(x) for x in re.findall(r"(\d+)º", ult)]
            if not nums:
                continue
            t_ini, n_t = min(nums), max(nums) - min(nums) + 1
            chave = [k for k, vv in G.NOME.items() if vv == disc_nome]
            if not chave:
                print(f"  [AVISO] {turma} sem{w} {G.DIAS[d]}: disciplina "
                      f"'{disc_nome}' não reconhecida, ignorada no relatório")
                continue
            disc_code = chave[0]
            doador = bool(G.posicoes_por_doador(turma, disc_code, d, t_ini, n_t))
            itens.append((w, d, t_ini, n_t, disc_code, prof_txt, doador))
    return itens


# ----------------------------------------------------------------- XLSX

def _linhas_coordenadas(comuns_por_par):
    linhas = []
    for (a, b), comuns in comuns_por_par.items():
        for disc, (tipo, _profs) in sorted(comuns.items()):
            situ = "Já combinada (grupo paralelo)" if tipo == "combinada" \
                else "Coordenada (professor comum, tempos diferentes)"
            linhas.append((f"{a}/{b}", G.NOME.get(disc, disc), situ))
    return linhas


def _linhas_trocas(itens_por_turma, comuns_por_par, irma, siglas):
    linhas = []
    for turma, itens in itens_por_turma.items():
        comuns = comuns_por_par.get(tuple(sorted((turma, irma.get(turma, turma)))), {})
        for (w, d, t, n, disc, prof, doador) in sorted(itens):
            obs = []
            if G.cruza_intervalo(t, n):
                obs.append("⚠ Cruza o intervalo do recreio — nenhuma outra "
                           "combinação coube")
            if disc in comuns:
                tipo, _ = comuns[disc]
                if tipo == "coordenar":
                    obs.append(f"Prova conjunta com {irma[turma]} (professor "
                               "comum, tempos coordenados)")
                else:
                    obs.append(f"Grupo paralelo já combinado com {irma[turma]}")
            obs_txt = "; ".join(obs)
            prof_nomes = nomes_dos_profs(prof, siglas)
            if not doador:
                if obs_txt:
                    linhas.append((turma, f"{G.NOME[disc]} / {prof_nomes}",
                                   "—", "—", "—", "—", obs_txt))
                continue
            for (dd, dp), tempos in G.posicoes_por_doador(turma, disc, d, t, n).items():
                tl = ", ".join(f"{x}º" for x in tempos) or f"{t + 1}º"
                dp_nome = nomes_dos_profs(dp, siglas)
                linhas.append((
                    turma, f"{G.NOME[disc]} / {prof_nomes}",
                    f"{tl} tempo(s) ({G.DIAS[d]}, semana {w})",
                    dp_nome, G.NOME.get(dd, dd),
                    f"Solicitar ao(à) prof(a). {dp_nome} a cessão do(s) "
                    f"tempo(s) {tl} de {G.NOME.get(dd, dd)} para a prova de "
                    f"{G.NOME[disc]}",
                    obs_txt))
    return linhas


def _linhas_regras_relaxadas(itens_por_turma, siglas):
    linhas = []
    for turma, itens in itens_por_turma.items():
        for (regra, dd, dp, detalhe) in G.detectar_regras_relaxadas(turma, itens):
            dp_nome = nomes_dos_profs(dp, siglas)
            linhas.append((turma, G.NOME_REGRA_CESSAO.get(regra, regra),
                           f"{G.NOME.get(dd, dd)} / {dp_nome}", detalhe))
    return sorted(linhas)


def _escrever_aba(wb, nome_aba, titulo, cabecalhos, linhas, larguras, col_esq=()):
    cab_fill = PatternFill("solid", fgColor="1F3864")
    cab_font = Font(bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    zebra = PatternFill("solid", fgColor="F2F2F2")
    n_col = len(cabecalhos)

    ws = wb.create_sheet(nome_aba)
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(n_col)}1")
    ws["A1"].alignment = Alignment(horizontal="center")

    for c, t in enumerate(cabecalhos, 1):
        cel = ws.cell(2, c, t)
        cel.fill, cel.font = cab_fill, cab_font
        cel.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        cel.border = borda

    i = 3
    for linha in linhas:
        for c, valor in enumerate(linha, 1):
            cel = ws.cell(i, c, valor)
            cel.border = borda
            cel.alignment = Alignment(
                vertical="center", wrap_text=(c in col_esq),
                horizontal="left" if c in col_esq else "center")
            if i % 2:
                cel.fill = zebra
        i += 1

    for col, larg in zip("ABCDEFGH", larguras):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A3"
    if linhas:
        ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(n_col)}{i - 1}"
    return ws


def exportar_xlsx(itens_por_turma, comuns_por_par, irma, siglas):
    destino = os.path.join(OUT, "Relatorio_trocas_de_tempo.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _escrever_aba(
        wb, "Provas coordenadas",
        "Provas coordenadas entre turmas irmãs — 2º semestre 2026 (Proposta 3)",
        ["Turmas", "Disciplina", "Situação"],
        _linhas_coordenadas(comuns_por_par),
        (14, 20, 42), col_esq=(3,))

    _escrever_aba(
        wb, "Trocas de tempo",
        "Trocas de tempo entre professores — 2º semestre 2026 (Proposta 3)",
        ["Turma", "Disciplina / Prof. solicitante", "Tempo necessário",
         "Prof. doador", "Disciplina do tempo doado", "Ação", "Observação"],
        _linhas_trocas(itens_por_turma, comuns_por_par, irma, siglas),
        (10, 38, 26, 30, 22, 60, 46), col_esq=(2, 4, 6, 7))

    _escrever_aba(
        wb, "Regras relaxadas",
        "Regras de cessão de aula relaxadas — 2º semestre 2026 (Proposta 3)",
        ["Turma", "Regra relaxada", "Disciplina / Professor", "Detalhe"],
        _linhas_regras_relaxadas(itens_por_turma, siglas),
        (10, 42, 34, 60), col_esq=(2, 3, 4))

    wb.save(destino)
    return destino


def main():
    G.carregar_ocupadas()
    caminho = f"{OUT}/Proposta_3_Calendario_Provas_2026_2SEM.xlsx"
    comuns_por_par = {(a, b): G.classificar_par(a, b) for a, b in G.PARES_IRMAS}
    itens_por_turma = {turma: ler_itens(caminho, turma) for turma in G.GRADES}

    p_md = G.relatorio({3: itens_por_turma}, comuns_por_par)
    print("Gerado:", p_md)

    irma = {a: b for a, b in G.PARES_IRMAS}
    irma.update({b: a for a, b in G.PARES_IRMAS})
    siglas = carregar_siglas()
    p_xlsx = exportar_xlsx(itens_por_turma, comuns_por_par, irma, siglas)
    print("Gerado:", p_xlsx)


if __name__ == "__main__":
    main()
